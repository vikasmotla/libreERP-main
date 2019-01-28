from __future__ import unicode_literals
from django.contrib.auth.models import User , Group
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate , login , logout
from django.contrib.auth.decorators import login_required
from django.core.urlresolvers import reverse
from django.template.loader import render_to_string, get_template
from django.template import RequestContext , Context
from django.conf import settings as globalSettings
from django.core.mail import send_mail , EmailMessage
from django.core import serializers
from django.http import HttpResponse ,StreamingHttpResponse
from django.utils import timezone
from django.db.models import Min , Sum , Avg , Q , F
import mimetypes
import hashlib, datetime, random
from datetime import timedelta , date
from monthdelta import monthdelta
from time import time
import pytz
import math
import json
from email.mime.image import MIMEImage
from bs4 import BeautifulSoup
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from StringIO import StringIO
import math
import requests
# related to the invoice generator
from reportlab import *
from reportlab.pdfgen import canvas
from reportlab.lib import colors, utils
from reportlab.platypus import Paragraph, Table, TableStyle, Image, Frame, Spacer, PageBreak, BaseDocTemplate, PageTemplate, SimpleDocTemplate, Flowable
from reportlab.platypus.flowables import HRFlowable
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet, TA_CENTER
from reportlab.graphics import barcode, renderPDF
# from reportlab.graphics.shapes import *
from reportlab.graphics.barcode.qr import QrCodeWidget
from reportlab.lib.pagesizes import letter, A5, A4, A3
from reportlab.lib.colors import *
from reportlab.lib.units import inch, cm ,mm
from reportlab.lib.enums import TA_JUSTIFY,TA_LEFT, TA_CENTER
from reportlab.graphics.barcode import code39
from reportlab.platypus.doctemplate import Indenter
from PIL import Image
from rest_framework import viewsets , permissions , serializers
from rest_framework.exceptions import *
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.renderers import JSONRenderer
from rest_framework.decorators import api_view
from url_filter.integrations.drf import DjangoFilterBackend
from .serializers import *
# from .helper import *
import datetime
from API.permissions import *
from HR.models import accountsKey
from reportlab.graphics.barcode.qr import QrCodeWidget
from django.core import serializers
from django.http import JsonResponse
from django.db.models import F ,Value,CharField,Prefetch
import ast
from reportlab.graphics import barcode , renderPDF
from django.utils import timezone

from django.shortcuts import render
from rest_framework import viewsets , permissions , serializers
from url_filter.integrations.drf import DjangoFilterBackend
from .serializers import *
from API.permissions import *
from .models import *
from HR.models import profile
from rest_framework.views import APIView
from rest_framework.renderers import JSONRenderer
from django.core.mail import send_mail, EmailMessage
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm, mm
from reportlab.lib import colors, utils
from reportlab.platypus import Paragraph, Table, TableStyle, Image, Frame, Spacer, PageBreak, BaseDocTemplate, PageTemplate, SimpleDocTemplate, Flowable
from reportlab.platypus.flowables import HRFlowable
from PIL import Image
from dateutil.relativedelta import relativedelta
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet, TA_CENTER
from reportlab.graphics import barcode, renderPDF
from reportlab.graphics.shapes import *
import calendar as pythonCal
from POS.models import *
from ERP.models import service, appSettingsField
from PIL import Image
from django.core.files.images import get_image_dimensions
import ast
from flask import Markup
from PIM.models import blogPost
from django.db.models.functions import Concat
from openpyxl import load_workbook
from io import BytesIO
import sendgrid
import os
from svglib.svglib import svg2rlg
from excel_response import ExcelResponse
import re
from ERP.send_email import send_email
from reportlab.platypus.flowables import Image as GNAA

# from sendgrid.helpers.mail import *

# Create your views here.

defaultSettingsData = appSettingsField.objects.filter(app_id=25)
fbLink = ''
lkLink = ''
twtLink = ''
if defaultSettingsData.count()>0:
    for i in defaultSettingsData:
        if i.name == 'facebookLink':
            fbLink = i.value
        elif i.name == 'twitterLink':
            lkLink = i.value
        elif i.name == 'linkedInLink':
            twtLink = i.value

from payu.gateway import get_hash , payu_url
from uuid import uuid4


def ecommerceHome(request):


    print 'home viewwwwwwwwwwwwwwwwwwwwwwww'
    data = {'wampServer' : globalSettings.WAMP_SERVER,'icon_logo':globalSettings.ICON_LOGO, 'useCDN' : globalSettings.USE_CDN,'brand_title':globalSettings.SEO_TITLE,'seoDetails':{'title':globalSettings.SEO_TITLE,'description':globalSettings.SEO_DESCRIPTION,'image':globalSettings.SEO_IMG,'width':globalSettings.SEO_IMG_WIDTH,'height':globalSettings.SEO_IMG_HEIGHT,'author':globalSettings.SEO_AUTHOR,'twitter_creator':globalSettings.SEO_TWITTER_CREATOR,'twitter_site':globalSettings.SEO_TWITTER_SITE,'site_name':globalSettings.SEO_SITE_NAME,'url':globalSettings.SEO_URL,'publisher':globalSettings.SEO_PUBLISHER} , 'color' : globalSettings.ECOMMERCE_THEME , "inventory" : globalSettings.INVENTORY_ENABLED , "settings" : application.objects.get(name = 'app.public.ecommerce' ).settings.all() }

    if '/' in request.get_full_path():
        urlData = request.get_full_path().split('/')
        print urlData,'url detailsssssssssss'
        if 'details'in urlData and len(urlData) > 2 :
            pk = urlData[-2]
            print pk
            try:
                lpObj = listing.objects.get(pk=pk)
                dpList = lpObj.files.all()
                print lpObj,'**************',dpList
                data['seoDetails']['title'] = lpObj.product.name
                data['seoDetails']['description'] = lpObj.product.description
                if len(dpList)>0:
                    try:
                        data['seoDetails']['image'] = dpList[0].attachment.url
                        w, h = get_image_dimensions(dpList[0].attachment.file)
                        print w,h
                        data['seoDetails']['width'] = w
                        data['seoDetails']['height'] = h
                    except:
                        print 'no such file has exists'
                    # image=Image.open(dpList[0].attachment.file)
                    # print image,image.size,image.format
            except:
                print 'please select the product'
        if 'blog' in urlData and len(urlData) > 1 :
            print 'blogggggggggggggggggggg'
            data['seoDetails']['title'] = str(globalSettings.SEO_TITLE) + ' | Blog'
        if 'categories' in urlData and len(urlData) > 2 :
            data['seoDetails']['title'] = str(urlData[-1]) + '| Buy ' + str(urlData[-1]) + ' At Best Price In India | ' + str(globalSettings.SEO_TITLE)
        if 'checkout' in urlData and len(urlData) > 2 :
            data['seoDetails']['title'] = str(globalSettings.SEO_TITLE) + ' | Review Order > Select Shipping Address > Place Order'
        if 'account' in urlData and len(urlData) > 2 and urlData[-1]!= '':
            print 'somethinggggggggggg'
            if urlData[-1] == 'cart':
                data['seoDetails']['title'] = str(globalSettings.SEO_TITLE) + ' | Shopping Cart'
            elif urlData[-1] == 'orders':
                data['seoDetails']['title'] = str(globalSettings.SEO_TITLE) + ' | My Orders'
            elif urlData[-1] == 'settings':
                data['seoDetails']['title'] = str(globalSettings.SEO_TITLE) + ' | My Settings'
            elif urlData[-1] == 'support':
                data['seoDetails']['title'] = str(globalSettings.SEO_TITLE) + ' | HelpCenter -  FAQ About Contextual Advertising , Online Advertising , Online Ads'
            elif urlData[-1] == 'saved':
                data['seoDetails']['title'] = str(globalSettings.SEO_TITLE) + ' | Saved Products'
        if len(urlData) > 1 :
            print 'pagessssssssssssss',urlData[1]
            pagesChecking = Pages.objects.filter(pageurl__icontains=str(urlData[1]))
            blogsChecking = blogPost.objects.filter(state__icontains='published',shortUrl__icontains=str(urlData[1]))
            if len(pagesChecking)>0:
                data['seoDetails']['title'] = str(globalSettings.SEO_TITLE) + ' |  ' + str(urlData[1]).replace('-',' ')
            elif len(blogsChecking)>0:
                blogData = blogsChecking[0]
                data['seoDetails']['title'] = str(globalSettings.SEO_TITLE) + ' |  ' + str(urlData[1]).replace('-',' ')
                if blogData.description is not None and len(blogData.description)>0 and blogData.description != 'null':
                    data['seoDetails']['description'] = blogData.description
                    print 'Desscription existsssssssssssss'
                else:
                    print 'noooooooooooo Desscription existsssssssssssss'
                try:
                    data['seoDetails']['image'] = blogData.ogimage.url
                    w, h = get_image_dimensions(blogData.ogimage.file)
                    print w,h
                    data['seoDetails']['width'] = w
                    data['seoDetails']['height'] = h
                    print 'og image existsssssssssssss'
                except:
                    print 'no such blog file has existsssssssssssss'
                    if blogData.ogimageUrl is not None and len(blogData.ogimageUrl)>0 and blogData.ogimageUrl != 'null':
                        data['seoDetails']['image'] = blogData.ogimageUrl
                    else:
                        print 'so ogimageurl existsssssssssssss'

    return render(request , 'ngEcommerce.html' , {'data':data})

class SearchProductAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny , )
    def get(self , request , format = None):
        if 'search' in self.request.GET:
            search = str(self.request.GET['search'])
            l = int(self.request.GET['limit'])
            print l , type(l)

            if 'multipleStore' in self.request.GET:
                print 'miltiiiiiiiiiiiiiiiiiiiiii',self.request.GET['pin']
                try:
                    store = Store.objects.get(pincode=int(self.request.GET['pin']))
                except:
                    return listing.objects.all()
                # storeQtylist = list(StoreQty.objects.filter(store = store.pk).values_list('pk',flat=True))
                # productsList = list(Product.objects.filter(storeQty__in=storeQtylist).values_list('pk',flat=True))
                productsList = list(StoreQty.objects.filter(store = store.pk).values_list('product',flat=True))
                # print productsList , 'productsList'
                listingobjs = listing.objects.filter(product__in=productsList)
                listingList = list(listingobjs.values_list('parentType',flat=True))
                genericList = genericProduct.objects.filter(pk__in=listingList)
                genericProd = list(genericList.filter(name__icontains=search).values('pk','name', 'visual').annotate(typ= Value('generic',output_field=CharField())))

                listProd = list(listingobjs.filter(Q(product__name__icontains=search) | Q(product__alias__icontains=search)).values('pk','product__id').annotate(name=F('product__name'), dp = F('files__attachment') ,inStock=F('product__inStock') , serialNo =F('product__serialNo'), howMuch =F('product__howMuch'),unit =F('product__unit') , dpId = F('files__imageIndex'), typ= Value('list',output_field=CharField())))
            else:
                # print globalSettings.INVENTORY_ENABLED

                genericProd = list(genericProduct.objects.filter(name__icontains=search).values('pk','name', 'visual').annotate(typ= Value('generic',output_field=CharField())))
                listProd = list(listing.objects.filter(Q(product__name__icontains=search) | Q(product__alias__icontains=search)).values('pk','product__id').annotate(name=F('product__name' ) , dp = F('files__attachment'), serialNo =F('product__serialNo'), howMuch =F('product__howMuch'), dpId = F('files__imageIndex') , unit = F('product__unit'), typ= Value('list',output_field=CharField())))
            newlist = []
            newListPks = []
            for i in listProd:
                if i['pk'] not in newListPks:
                    if not globalSettings.INVENTORY_ENABLED:
                        i['inStock'] = 1000
                    else:
                        storeQtyList = StoreQty.objects.filter(product__id = i['product__id'] ,productVariant__isnull = True)
                        if len(storeQtyList)>0:
                            i['inStock'] = storeQtyList[0].quantity
                        else:
                            i['inStock'] = 0
                    print i['inStock']
                    newlist.append(i)
                    prodVar = ProductVerient.objects.filter(parent = i['product__id'])
                    if len(prodVar)>0:
                        print 'prod varienttttttttttt'
                        for a in prodVar:
                            prodVar = {'name':i['name'] , 'dp':i['dp'], 'serialNo':a.sku,'prodDesc':a.prodDesc ,'unitPerpack':a.unitPerpack ,'howMuch':i['howMuch']* a.unitPerpack , 'unit':i['unit'] ,'typ':i['typ'] ,'pk':i['pk'] }
                            if not globalSettings.INVENTORY_ENABLED:
                                prodVar['inStock'] = 1000
                            else:
                                storeQtyList = StoreQty.objects.filter(product__id = i['product__id'] ,productVariant__id = a.pk)
                                if len(storeQtyList)>0:
                                    print 'yesssssssssssssss'
                                    prodVar['inStock'] = storeQtyList[0].quantity
                                else:
                                    print 'nonnnnnnnnnnnn'
                                    prodVar['inStock'] = 0
                            newlist.append(prodVar)
                    newListPks.append(i['pk'])
            tosend = genericProd + newlist
            return Response(tosend[0:l], status = status.HTTP_200_OK)

class PromoCheckAPI(APIView):
    renderer_classes = (JSONRenderer,)
    # permission_classes = (permissions.IsAuthenticated ,)
    def get(self , request , format = None):
        if 'name' in self.request.GET:
            print self.request.GET['name']
            promObj = Promocode.objects.filter(name = self.request.GET['name'])
            val = 0
            if len(promObj)>0:
                if timezone.now()<=promObj[0].endDate:
                    orderCount = Order.objects.filter(~Q(status='failed'),user=request.user,promoCode=self.request.GET['name']).count()
                    toReturn = 'Success' if orderCount<promObj[0].validTimes else 'Already Used'
                    val = promObj[0].discount if toReturn=='Success' else 0
                else:
                    toReturn = 'Promocode Has Expired'
            else:
                toReturn = 'Invalid Promocode'
            return Response({'msg':toReturn,'val':val}, status = status.HTTP_200_OK)

from rate_request import getShipmentCharges
class ShipmentChargeAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def get(self , request , format = None):
        print 'hereeeeee'
        return Response(getShipmentCharges(request.GET['pincode'] , request.GET['country'] , float(request.GET['weight'])    ))

class CategorySortListAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def get(self , request , format = None):
        gpList = list(genericProduct.objects.filter(parent__isnull=True).values('name','pk','restricted').annotate(img=Concat(Value('/media/'),'visual')))
        print gpList,'GPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPP#####################33'
        user = request.user
        for idx, val in enumerate(gpList):
            if val['restricted']==True:
                if user.is_staff==True:
                    chilData = list(genericProduct.objects.filter(parent=val['pk']).values('name','pk').annotate(img=Concat(Value('/media/'),'visual')))
                    if len(chilData)>0:
                        val['child'] = chilData
                    else:
                        val['child'] = []
                else:
                    del gpList[idx]
            else:
                chilData = list(genericProduct.objects.filter(parent=val['pk']).values('name','pk').annotate(img=Concat(Value('/media/'),'visual')))
                if len(chilData)>0:
                    val['child'] = chilData
                else:
                    val['child'] = []

        return Response(gpList, status = status.HTTP_200_OK)


class CreateOrderAPI(APIView):
    renderer_classes = (JSONRenderer,)
    # permission_classes = (permissions.IsAuthenticated ,)
    def post(self , request , format = None):
        print request.data
        oQMp = []
        totalAmount = 0
        msg = 'Error'
        contactData=[]
        userCart = Cart.objects.filter(user=request.user)
        # print userCart.count(),userCart
        # print request.data,type(request.data),'fffffffffffffffff'
        prod = request.data['products']
        if type(request.data['products']) == unicode:
            prod = ast.literal_eval(request.data['products'])
        print 'ssssssssssssssss',prod,type(prod)
        orderGst = 0
        pdAmount = 0
        for i in prod:
            pObj = listing.objects.get(pk = i['pk'])
            qtyGst = 0
            taxAmount = 0
            isStoreGlobal = True
            globalStoreObj = appSettingsField.objects.filter(name='isStoreGlobal')
            if len(globalStoreObj)>0:
                if globalStoreObj[0].flag:
                    isStoreGlobal = False

            if pObj.product.serialNo == i['prodSku']:
                pp = pObj.product.price
                if pObj.product.productMeta:
                    ptaxRate = pObj.product.productMeta.taxRate
                    taxAmount = float(round((pp * ptaxRate)/100 , 2))
                if pp > 0:
                    a = pp - (pObj.product.discount*pp)/100
                    # print a ,type(a)
                    b = a - float(request.data['promoCodeDiscount'])*a/100
                    finalTaxAmount = taxAmount - float(request.data['promoCodeDiscount']*taxAmount/100)
                else:
                    b=0
                    finalTaxAmount = 0
                priceDuringOrder = pp
            else:
                prodVar = ProductVerient.objects.get(sku = i['prodSku'])
                pp = prodVar.discountedPrice
                priceDuringOrder = pp
                if prodVar.parent.productMeta:
                    ptaxRate = prodVar.parent.productMeta.taxRate
                    taxAmount = float(round((pp * ptaxRate)/100, 2))
                if pp > 0:
                    b = pp - float(request.data['promoCodeDiscount'])*pp/100
                    finalTaxAmount = taxAmount - float(request.data['promoCodeDiscount'])*taxAmount/100
                else:
                    b=0
                    finalTaxAmount = 0

            pDiscAmount = round(pp-b, 2)
            if isStoreGlobal:
                pDiscAmount += round(taxAmount-finalTaxAmount, 2)
                b += finalTaxAmount
            else:
                pass

            pDiscAmount = pDiscAmount * i['qty']
            qtyGst = finalTaxAmount * i['qty']
            orderGst += qtyGst
            qtyTotal = b * i['qty']
            totalAmount += qtyTotal

            oQMObj = OrderQtyMap.objects.create(**{'product':pObj,'qty':i['qty'],'totalAmount':round(qtyTotal,2),'discountAmount':round(pDiscAmount, 2),'prodSku':i['prodSku'],'desc':i['desc'],'gstAmount':round(qtyGst, 2),'priceDuringOrder':round(priceDuringOrder, 2)})
            oQMp.append(oQMObj)
            obj = pObj.product
            if 'storepk' in request.data:
                print 'multistore'
                storeObj = Store.objects.get(pk=int(request.data['storepk']))
                try:
                    if i['prodSku']==obj.serialId:
                        print 'parent, multistoreeeeeeeeeeeeeeeeee',storeObj.name
                        storeQtyObj = StoreQty.objects.get(store__id = storeObj.pk, product__id = obj.pk, productVariant__isnull = True)
                        if i['qty']>storeQtyObj.quantity:
                            print 'error raise exception'
                        storeQtyObj.quantity = storeQtyObj.quantity - i['qty']
                        storeQtyObj.save()
                    else:
                        print 'child, multistoreeeeeeeeeeeeeeeeee'
                        prodVar = ProductVerient.objects.get(parent__id = obj.pk , sku = i['prodSku'])
                        storeQtyObj = StoreQty.objects.get(store__id = storeObj.pk, product__id = obj.pk, productVariant__id = prodVar.pk)
                        if i['qty']>storeQtyObj.quantity:
                            print 'error raise exception'
                        storeQtyObj.quantity = storeQtyObj.quantity - i['qty']
                        storeQtyObj.save()
                except :
                    print 'no deduction in qantity , item doesnt exist in storeQty'

            else:
                try:
                    if i['prodSku']==obj.serialId:
                        print 'parent , single store'
                        storeQtyObj = StoreQty.objects.get(master = True , product__id = obj.pk, productVariant__isnull = True)
                        if i['qty']>storeQtyObj.quantity:
                            print 'error raise exception'
                        storeQtyObj.quantity = storeQtyObj.quantity - i['qty']
                        storeQtyObj.save()
                    else:
                        print 'child , single store'
                        prodVar = ProductVerient.objects.get(parent__id = obj.pk , sku = i['prodSku'])
                        storeQtyObj = StoreQty.objects.get(master = True, product__id = obj.pk, productVariant__id = prodVar.pk)
                        if i['qty']>storeQtyObj.quantity:
                            print 'error raise exception'
                        storeQtyObj.quantity = storeQtyObj.quantity - i['qty']
                        storeQtyObj.save()
                except :
                    print 'no deduction in qantity , item doesnt exist in storeQty'

        else:
            data = {
            'user':User.objects.get(pk=request.user.pk),
            'totalAmount' : round(totalAmount,2),
            'paymentMode' : str(request.data['modeOfPayment']),
            'modeOfShopping' : str(request.data['modeOfShopping']),
            'totalGst' : orderGst,
            'landMark' : str(request.data['address']['landMark']),
            'street' : str(request.data['address']['street']),
            'city' : str(request.data['address']['city']),
            'state' : str(request.data['address']['state']),
            'pincode' : str(request.data['address']['pincode']),
            'country' : str(request.data['address']['country']),
            'mobileNo' : str(request.data['address']['mobileNo']),
            'billingLandMark' : str(request.data['billingAddress']['landMark']),
            'billingStreet' : str(request.data['billingAddress']['street']),
            'billingCity' : str(request.data['billingAddress']['city']),
            'billingState' : str(request.data['billingAddress']['state']),
            'billingPincode' : str(request.data['billingAddress']['pincode']),
            'billingCountry' : str(request.data['billingAddress']['country']),
            }

            try:
                countryCode = Countries.objects.filter(name__iexact = str(request.data['address']['country']))
                data['countryCode'] = countryCode[0].sortname
            except :
                pass

            if len(str(request.data['promoCode'])) > 0:
                data['promoCode'] = str(request.data['promoCode'])
            print data
            if 'shippingCharges' in request.data:
                shippingCharges = round(request.data['shippingCharges'],2)
                data['shippingCharges'] = shippingCharges
                data['totalAmount'] += shippingCharges
            orderObj = Order.objects.create(**data)
            for i in oQMp:
                orderObj.orderQtyMap.add(i)
            orderObj.save()
            msg = 'Sucess'
            if str(request.data['modeOfPayment']) == 'COD':
                userCart.delete()
            # response = HttpResponse(content_type='application/pdf')
            # response['Content-Disposition'] = 'attachment; filename="order_invoice%s_%s_%s.pdf"' % (
            # orderObj.totalAmount, datetime.datetime.now(pytz.timezone('Asia/Kolkata')).year, orderObj.pk)
            # genInvoice(response, orderObj, request)
            # f = open(os.path.join(globalSettings.BASE_DIR, 'media_root/order_invoice%s%s_%s.pdf' %
            #                   ( orderObj.totalAmount, datetime.datetime.now(pytz.timezone('Asia/Kolkata')).year, orderObj.pk)), 'wb')
            # f.write(response.content)
            # f.close()
            value = []
            totalPrice = 0
            promoAmount = 0
            total=0
            price = 0
            grandTotal = 0
            promoObj = Promocode.objects.all()
            for p in promoObj:
                if str(p.name)==str(orderObj.promoCode):
                    promoAmount = p.discount
            print promoAmount
            a = '#'
            docID = str(a) + str(orderObj.pk)
            for i in orderObj.orderQtyMap.all():
                if i.prodSku == i.product.product.serialNo:
                    price = i.product.product.price - (i.product.product.discount * i.product.product.price)/100
                    price=round(price, 2)
                    totalPrice=i.qty*price
                    totalPrice=round(totalPrice, 2)
                    total+=totalPrice
                    total=round(total, 2)
                    qtyData =  i.product.product.howMuch
                    if str(i.product.product.unit)=='Gram' or str(i.product.product.unit)=='gm':
                        if qtyData >1000:
                            qtyValue = str(qtyData/1000) + ' Kg'
                        else:
                            qtyValue = str(qtyData) + ' gm'
                    elif str(i.product.product.unit)=='Millilitre' or str(i.product.product.unit)=='ml':
                        if qtyData>1000:
                            qtyValue = str(qtyData/1000) + ' lt'
                        else:
                            qtyValue = str(qtyData) + ' ml'
                    elif str(i.product.product.unit)=='Size and Color' or str(i.product.product.unit)=='Size':
                        if int(qtyData)==1:
                            qtyValue = 'XS'
                        elif int(qtyData)==2:
                            qtyValue = 'S'
                        elif int(qtyData)==2:
                            qtyValue = 'M'
                        elif int(qtyData)==4:
                            qtyValue = 'L'
                        elif int(qtyData)==5:
                            qtyValue = 'XL'
                        elif int(qtyData)==6:
                            qtyValue = 'XL'
                        else:
                            qtyValue = qtyData
                    else:
                      qtyValue = qtyData

                    if i.desc:
                         desc = i.desc
                    else:
                        desc =""
                    productName = str(i.product.product.name) + ' ' + str(qtyValue)+ ' ' + str(desc)
                    value.append({ "productName" : productName,"qty" : i.qty , "amount" : i.totalAmount,"price":price , "gst":i.gstAmount})
                else:
                    prodData = ProductVerient.objects.get(sku = i.prodSku)
                    price = prodData.discountedPrice
                    price=round(price, 2)
                    totalPrice=i.qty*price
                    totalPrice=round(totalPrice, 2)
                    total+=totalPrice
                    total=round(total, 2)
                    qtyData = prodData.unitPerpack * i.product.product.howMuch
                    if str(i.product.product.unit)=='Gram' or str(i.product.product.unit)=='gm':
                        if qtyData >1000:
                            qtyValue = str(qtyData/1000) + ' Kg'
                        else:
                            qtyValue = str(qtyData) + ' gm'
                    elif str(i.product.product.unit)=='Millilitre' or str(i.product.product.unit)=='ml':
                        if qtyData>1000:
                            qtyValue = str(qtyData/1000) + ' lt'
                        else:
                            qtyValue = str(qtyData) + ' ml'
                    elif str(i.product.product.unit)=='Size and Color' or str(i.product.product.unit)=='Size':
                        if int(qtyData)==1:
                            qtyValue = 'XS'
                        elif int(qtyData)==2:
                            qtyValue = 'S'
                        elif int(qtyData)==2:
                            qtyValue = 'M'
                        elif int(qtyData)==4:
                            qtyValue = 'L'
                        elif int(qtyData)==5:
                            qtyValue = 'XL'
                        elif int(qtyData)==6:
                            qtyValue = 'XL'
                        else:
                            qtyValue = qtyData
                    else:
                      qtyValue = qtyData

                    print qtyValue ,'ddddddddddddddddddddddddddddddddd'
                    if i.desc:
                         desc = i.desc
                    else:
                        desc =""
                    productName = str(i.product.product.name) + ' ' + str(qtyValue)+ ' ' + str(desc)
                    value.append({ "productName" : productName,"qty" : i.qty , "amount" : i.totalAmount,"price":price, "gst":i.gstAmount})

            shippingCharges = round(orderObj.shippingCharges,2)
            grandTotal = round(orderObj.totalAmount, 2)
            try:
                isStoreGlobal = appSettingsField.objects.filter(name='isStoreGlobal')[0].flag
            except:
                isStoreGlobal = False
            try:
                companyAddress = appSettingsField.objects.filter(name='companyAddress')[0].value
                cleanr = re.compile('<.*?>')
                companyAddress = re.sub(cleanr, '', companyAddress)
                print companyAddress ,'&****************************************************'
            except:
                companyAddress = ''
                print companyAddress ,'@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@2'
            try:
                gstValue = appSettingsField.objects.filter(name='cstNo')[0].value
            except:
                gstValue = ''
            try:
                currencyTyp = appSettingsField.objects.filter(name='currencySymbol')[0].value
                if currencyTyp == 'fa-usd':
                    currencyVal = 'USD'
                else:
                    currencyVal = 'INR'
            except:
                currencyVal = ''
            if orderObj.user.email and orderObj.paymentMode == 'COD':
                ctx = {
                    'heading' : "Invoice Details",
                    'recieverName' : orderObj.user.first_name  + " " +orderObj.user.last_name ,
                    'linkUrl': globalSettings.BRAND_NAME,
                    'sendersAddress' : globalSettings.SEO_TITLE,
                    # 'sendersPhone' : '122004',
                    'grandTotal':grandTotal,
                    'total': total,
                    'value':value,
                    'docID':docID,
                    'data':orderObj,
                    'promoAmount':promoAmount,
                    'linkedinUrl' : lkLink,
                    'fbUrl' : fbLink,
                    'twitterUrl' : twtLink,
                    'isStoreGlobal':isStoreGlobal,
                    'companyAddress':companyAddress,
                    'gstValue':gstValue,
                    'currencyVal':currencyVal,
                    'shippingCharges':shippingCharges
                }
                print ctx , 'ctxxxxxxxxxxxx'
                email_body = get_template('app.ecommerce.emailDetail.html').render(ctx)
                email_subject = "Order Details"
                email_to = []
                email_to.append(str(orderObj.user.email))
                for i in globalSettings.G_ADMIN:
                    email_to.append(str(i))
                email_cc = []
                email_bcc = []
                send_email(email_body,email_to,email_subject,email_cc,email_bcc,'html')
            return Response({'paymentMode':orderObj.paymentMode,'dt':orderObj.created,'odnumber':orderObj.pk}, status = status.HTTP_200_OK)


class fieldViewSet(viewsets.ModelViewSet):
    permission_classes = (isAdmin , )
    queryset = field.objects.all()
    serializer_class = fieldSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['name']

class genericProductViewSet(viewsets.ModelViewSet):
    permission_classes = (isAdminOrReadOnly , )
    # queryset = genericProduct.objects.all()
    serializer_class = genericProductSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['name','parent']
    def get_queryset(self):
        if 'multipleStore' in self.request.GET:
            print 'miltiiiiiiiiiiiiiiiiiiiiii',self.request.GET['pin']
            try:
                store = Store.objects.get(pincode=int(self.request.GET['pin']))
            except:
                return listing.objects.all()
            # storeQtylist = list(StoreQty.objects.filter(store = store.pk).values_list('pk',flat=True))
            # productsList = list(Product.objects.filter(storeQty__in=storeQtylist).values_list('pk',flat=True))
            productsList = list(StoreQty.objects.filter(store = store.pk).values_list('product',flat=True))
            listingList = list(listing.objects.filter(product__in=productsList).values_list('parentType',flat=True))
            genericList = genericProduct.objects.filter(pk__in=listingList)
            if 'genericValue' in  self.request.GET:
                return  genericList.exclude(pk = self.request.GET['genericValue'])[0:4]
            return genericList
        if 'genericValue' in  self.request.GET:
            return  genericProduct.objects.exclude(pk = self.request.GET['genericValue'] ).filter()[0:4]
        else:
            return genericProduct.objects.all()


class mediaViewSet(viewsets.ModelViewSet):
    permission_classes = (isAdminOrReadOnly , )
    queryset = media.objects.all()
    serializer_class = mediaSerializer


def getProducts(lst , node , multiproductLst):
    a = multiproductLst
    lst = lst | listing.objects.filter(parentType = node,product__in=a)
    for child in node.children.all():
        lst = getProducts(lst , child , a)
    return lst

class listingViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny , )
    serializer_class = listingSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['parentType','product']

    def get_queryset(self):

        # abc =  listing.objects.all()
        # for i in abc:
        #     print i.files.order_by('imageIndex')


        #     for im in imagelist:
        #         print im.imageIndex , im.pk ,'uuuuuuuurrrrrrrrrrrrrllllllllllllllllll'
        # print abc

        print self.request.GET,'aaaaaaaaaaaaaaaaaaaaaaaaaaaaaa'
        data = self.request.GET
        if 'recursive' in data:
            if data['recursive'] == '1':
                print 'enterrrrrrrrr'
                print data['parent'] , type(data['parent'])
                prnt = genericProduct.objects.get(id = data['parent'])
                print prnt
                toReturn = listing.objects.filter(parentType = prnt).order_by(F('productIndex').asc())
                multiproductLst = []
                if 'multipleStore' in self.request.GET:
                    print 'miltiiiiiiiiiiiiiiiiiiiiii',self.request.GET['pin']
                    try:
                        store = Store.objects.get(pincode=int(self.request.GET['pin']))
                    except:
                        return listing.objects.all()
                    # storeQtylist = list(StoreQty.objects.filter(store = store.pk).values_list('pk',flat=True))
                    # productsList = list(Product.objects.filter(storeQty__in=storeQtylist).values_list('pk',flat=True))
                    productsList = list(StoreQty.objects.filter(store = store.pk).values_list('product',flat=True))
                    toReturn = toReturn.filter(product__in=productsList)
                    multiproductLst = productsList

                print toReturn.count()

                for child in prnt.children.all():
                    toReturn = getProducts(toReturn , child , multiproductLst)

                if 'minPrice' in data:
                    minPrice = data['minPrice']
                    maxPrice = data['maxPrice']
                    toReturn = toReturn.filter(product__price__lte = maxPrice , product__price__gte = minPrice)

                if 'fields' in data:
                    d = ast.literal_eval(data['fields'])
                    if len(d)!=0:
                        count = 0
                        for k,v in d.items():
                            if count == 0:
                                count += 1
                                if type(v)==list:
                                    for idx,c in enumerate(v):
                                        if idx == 0:
                                            qry1 = Q(specifications__icontains = '"name":"{0}","value":"{1}"'.format(k,c))
                                        else:
                                            qry1 = qry1 | Q(specifications__icontains = '"name":"{0}","value":"{1}"'.format(k,c))
                                    qry = qry1
                                else:
                                    # qry = Q(dfs__name=k,dfs__value=v)
                                    qry = Q(specifications__icontains = '"name":"{0}","value":"{1}"'.format(k,v))
                            else:
                                if type(v)==list:
                                    for idx,c in enumerate(v):
                                        if idx == 0:
                                            qry2 = Q(specifications__icontains = '"name":"{0}","value":"{1}"'.format(k,c))
                                        else:
                                            qry2 = qry2 | Q(specifications__icontains = '"name":"{0}","value":"{1}"'.format(k,c))
                                    qry = qry & qry2
                                else:
                                    # qry = qry & Q(dfs__name=k,dfs__value=v)
                                    qry = qry & Q(specifications__icontains = '"name":"{0}","value":"{1}"'.format(k,v))
                        print 'gggggggggggqqqqqqqqqqq',qry
                        toReturn = toReturn.filter(qry)
                        print 'filtered',toReturn
                        print len(toReturn)

                if 'sort' in data:
                    if data['sort']=='lth':
                        toReturn = toReturn.order_by('product__price')
                    elif data['sort']=='htl':
                        toReturn = toReturn.order_by('-product__price')
                # for idx,c in enumerate(self.request.GET[]):
                #         if idx == 0:
                #             qry = Q(specifications__icontains = '"name":"place","value":"'+ cities[idx])
                #         else:
                #             qry = qry | Q(specifications__icontains = '"name":"place","value":"'+ cities[idx])

                # if 'city' in self.request.GET:
                #     cities = self.request.GET.getlist('city')
                #     cL = len(cities)
                #     for idx,c in enumerate(cities):
                #         if idx == 0:
                #             qry = Q(specifications__icontains = '"name":"place","value":"'+ cities[idx])
                #         else:
                #             qry = qry | Q(specifications__icontains = '"name":"place","value":"'+ cities[idx])

                    # toReturn = toReturn.filter(qry)
                # print toReturn
                # for i in toReturn:
                #     print i.product
                #
                #     product  = Product.objects.filter(pk__in=i.product)
                #     print product ,'aaaaaaaaaaaaaa'
                #     prductSku =  ProductVerient.objects.all()
                #     for i in product:
                #         print i,'jjjjj'
                #         for j in prductSku:
                #             print j.parent,'bbbbbbbbbbbbbbbb'
                #             if i == j.parent:
                #                 print 'kkkkkkkkkkkkkkkkkk'
                return toReturn
        else:
            # return listing.objects.all()
            return listing.objects.all()


class listingLiteViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny, )
    serializer_class = listingLiteSerializer
    # queryset = listing.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['product','parentType']
    def get_queryset(self):
        # if self.request.user.is_authenticated:
        #     print "gggggggggggggggggggggg", self.request.user
        #     u = self.request.user
        #     has_application_permission(u , ['app.ecommerce' , 'app.ecommerce.listings'])
        if 'multipleStore' in self.request.GET:
            print 'miltiiiiiiiiiiiiiiiiiiiiii',self.request.GET['pin']
            try:
                store = Store.objects.get(pincode=int(self.request.GET['pin']))
            except:
                return listing.objects.all()
            # storeQtylist = list(StoreQty.objects.filter(store = store.pk).values_list('pk',flat=True))
            # productsList = list(Product.objects.filter(storeQty__in=storeQtylist).values_list('pk',flat=True))
            productsList = list(StoreQty.objects.filter(store = store.pk).values_list('product',flat=True))
            listingList = listing.objects.filter(product__in=productsList)
            if 'parentValue' in  self.request.GET:
                return listingList.filter(parentType=self.request.GET['parentValue']).exclude(pk = self.request.GET['detailValue'] )[0:4]
            return listingList

        if 'parentValue' in  self.request.GET:
            return listing.objects.filter(parentType=self.request.GET['parentValue']).exclude(pk = self.request.GET['detailValue'] )[0:4]
        # else:
        #     return listing.objects.all()
        if 'mode' in  self.request.GET:
            if self.request.GET['mode'] == 'vendor':
                s = service.objects.get(user = u)
                items = offering.objects.filter( service = s).values_list('item' , flat = True)
                return listing.objects.exclude(pk__in = items)
            elif self.request.GET['mode'] == 'suggest':
                return listing.objects.all()[:5]
        # else:
        return listing.objects.all()
        # return listing.objects.all().order_by(F('productIndex').asc())
        #     if self.request.GET['parentValue'] == 'vendor':
        #         s = service.objects.get(user = u)
        #         items = offering.objects.filter( service = s).values_list('item' , flat = True)
        #         return listing.objects.exclude(pk__in = items)
        #     elif self.request.GET['mode'] == 'suggest':
        #         return listing.objects.all()[:5]
        # else:
        #     return listing.objects.all()


class CategoryViewSet(viewsets.ModelViewSet):
    permission_classes = (isAdminOrReadOnly , )
    queryset = Category.objects.all()
    serializer_class = CategorySerializer

class PagesViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny ,)
    queryset = Pages.objects.all()
    serializer_class = PagesSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['title','pageurl' , 'topLevelMenu']

class offerBannerViewSet(viewsets.ModelViewSet):
    permission_classes = (isAdminOrReadOnly, )
    serializer_class = offerBannerSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['title','level']
    def get_queryset(self):
        return offerBanner.objects.all()
        # return offerBanner.objects.filter(active = True)

class CartViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny , )
    serializer_class = CartSerializer
    queryset = Cart.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['user','typ','product']


class ActivitiesViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny , )
    serializer_class = ActivitiesSerializer
    # queryset = Activities.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['user','typ','product']
    def get_queryset(self):
        a = Activities.objects.filter(user=self.request.user,product__isnull=False).order_by('-created')
        if 'multipleStore' in self.request.GET:
            print 'miltiiiiiiiiiiiiiiiiiiiiii',self.request.GET['pin']
            try:
                store = Store.objects.get(pincode=int(self.request.GET['pin']))
            except:
                return listing.objects.all()
            # storeQtylist = list(StoreQty.objects.filter(store = store.pk).values_list('pk',flat=True))
            # productsList = list(Product.objects.filter(storeQty__in=storeQtylist).values_list('pk',flat=True))
            productsList = list(StoreQty.objects.filter(store = store.pk).values_list('product',flat=True))
            listingList = list(listing.objects.filter(product__in=productsList).values_list('pk',flat=True))
            a = a.filter(product__in = listingList)
        pPk = []
        aPk = []
        for i in a :
            if len(aPk) == 4:
                break
            if i.product.pk not in pPk:
                pPk.append(i.product.pk)
                aPk.append(i.pk)
            else:
                a = a.exclude(product=i.product)
        return Activities.objects.filter(pk__in=aPk).order_by('-created')

class AddressViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny , )
    queryset = Address.objects.all()
    serializer_class = AddressSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['user','pincode']

class TrackingLogViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny , )
    queryset = TrackingLog.objects.all()
    serializer_class = TrackingLogSerializer

class OrderQtyMapViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny , )
    queryset = OrderQtyMap.objects.all()
    serializer_class = OrderQtyMapSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['courierAWBNo','orderBy']

class OrderViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny , )
    # queryset = Order.objects.all()
    serializer_class = OrderSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['status','id','orderQtyMap']
    def get_queryset(self):
        # return Order.objects.filter( ~Q(status = 'failed')).order_by('-created')
        toRet = Order.objects.all().order_by('-created')
        if 'deliveryCenter_filter' in self.request.GET:
            sts = str(self.request.GET['deliveryCenter_filter'])
            toRet = toRet.filter(orderQtyMap__status=sts).distinct()
        if 'user' in self.request.GET:
            return toRet.filter(user=self.request.user)
        else:
            return toRet

class PromocodeViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny , )
    queryset = Promocode.objects.all()
    serializer_class = PromocodeSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['name']

class FrequentlyQuestionsViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny , )
    queryset = FrequentlyQuestions.objects.all()
    serializer_class = FrequentlyQuestionsSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['ques']

class PincodeViewSet(viewsets.ModelViewSet):
    permission_classes = (isAdminOrReadOnly , )
    queryset = Pincode.objects.all()
    serializer_class = pincodeSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['pincodes']



def manifest(response,item,typ,packingSlip):
    print packingSlip ,'hhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhh'
    settingsFields = application.objects.get(name = 'app.public.ecommerce').settings.all()
    # print settingsFields.get(name = 'companyAddress').value

    now = datetime.datetime.now()
    # print now.year,now.month,now.day

    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(response,pagesize=letter, topMargin=1*cm,leftMargin=0.2*cm,rightMargin=0.2*cm)
    elements = []
    if typ == 'all':
        order = item
        orderQtsObj = item.orderQtyMap.all()
        courierName = orderQtsObj[0].courierName
        courierAWBNo = orderQtsObj[0].courierAWBNo
        total = 0
        for i in orderQtsObj:
            total += (i.totalAmount-i.discountAmount) * i.qty
    else:
        order = item.order.get()
        orderQtsObj = [item]
        courierName = item.courierName
        courierAWBNo = item.courierAWBNo
        total = (item.totalAmount-item.discountAmount) * item.qty

    currencySymbol = appSettingsField.objects.filter(name='currencySymbol')
    if len(currencySymbol)>0:
        if currencySymbol[0].value == 'fa-usd':
            currency = '(USD)'
        else:
            currency = '(INR)'
    else:
        currency = '(USD)'

    elements.append(HRFlowable(width="100%", thickness=1, color=black,spaceAfter=10))

    if not packingSlip:
        if order.paymentMode == 'card':
            txt1 = '<para size=13 leftIndent=150 rightIndent=150><b>PREPAID - DO NOT COLLECT CASH</b></para>'
        else:
            txt1 = '<para size=13 leftIndent=150 rightIndent=150><b>CASH ON DELIVERY &nbsp; {0} INR</b></para>'.format(total)
        elements.append(Paragraph(txt1, styles['Normal']))
        elements.append(Spacer(1, 8))
    txt2 = '<para size=10 leftIndent=150 rightIndent=150><b>DELIVERY ADDRESS :</b> {0},<br/>{1},<br/>{2} - {3},<br/>{4} , {5}.</para>'.format(order.landMark,order.street,order.city,order.pincode,order.state,order.country)
    elements.append(Paragraph(txt2, styles['Normal']))
    elements.append(Spacer(1, 30))

    txt3 = '<para size=10 leftIndent=150 rightIndent=150><b>COURIER NAME : </b>{0}<br/><b>COURIER AWB No. : </b>{1}</para>'.format(courierName,courierAWBNo)
    elements.append(Paragraph(txt3, styles['Normal']))
    elements.append(Spacer(1, 10))

    txt4 = '<para size=10 leftIndent=150 rightIndent=150><b>SOLD BY : </b>{0}</para>'.format(settingsFields.get(name = 'companyAddress').value)
    elements.append(Paragraph(txt4, styles['Normal']))
    elements.append(Spacer(1, 3))
    # txt5 = '<para size=10 leftIndent=150 rightIndent=150><b>VAT/TIN No. : </b>{0} &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;<br/><b>CST No. : </b>{1}</para>'.format(settingsFields.get(name = 'vat_tinNo').value,settingsFields.get(name = 'cstNo').value)
    txt5 = '<para size=10 leftIndent=150 rightIndent=150><b>GST No. : </b>{0}</para>'.format(settingsFields.get(name = 'cstNo').value)
    elements.append(Paragraph(txt5, styles['Normal']))
    elements.append(Spacer(1, 10))
    invNo = str(now.year)+str(now.month)+str(now.day)+str(order.pk)
    txt6 = '<para size=10 leftIndent=150 rightIndent=150><b>Invoice No. : </b>{0} </para>'.format(invNo)
    elements.append(Paragraph(txt6, styles['Normal']))
    elements.append(Spacer(1, 30))
    tableDataPackingSlip = [['SKU','Product','Qty']]
    tableData=[['Product','Price','Qty','Discount','Final Price']]
    for item in orderQtsObj:
        # print item,item.status,'rrrrrrrrrrrrrrrrrrrrrrrr'
        if item.status == 'created':
            item.status = 'packed'
            item.save()
        if item.prodSku != item.product.product.serialNo:
            product = ProductVerient.objects.get(sku=item.prodSku)
            qtyData = product.unitPerpack * item.product.product.howMuch
        else:
            qtyData =  item.product.product.howMuch

        if str(item.product.product.unit)=='Gram' or str(item.product.product.unit)=='gm':
            if qtyData >1000:
                qtyValue = str(qtyData/1000) + ' Kg'
            else:
                qtyValue = str(qtyData) + ' gm'
        elif str(item.product.product.unit)=='Millilitre' or str(item.product.product.unit)=='ml':
            if qtyData>1000:
                qtyValue = str(qtyData/1000) + ' lt'
            else:
                qtyValue = str(qtyData) + ' ml'
        elif str(item.product.product.unit)=='Size and Color' or str(item.product.product.unit)=='Size':
            if int(qtyData)==1:
                qtyValue = 'XS'
            elif int(qtyData)==2:
                qtyValue = 'S'
            elif int(qtyData)==2:
                qtyValue = 'M'
            elif int(qtyData)==4:
                qtyValue = 'L'
            elif int(qtyData)==5:
                qtyValue = 'XL'
            elif int(qtyData)==6:
                qtyValue = 'XL'
            else:
                qtyValue = qtyData
        else:
          qtyValue = qtyData

        if item.desc:
             desc = item.desc
        else:
            desc =""

        name = str(item.product.product.name)+ ' ' + str(qtyValue)+ ' ' +str(desc)
        sku = str(item.product.product.serialNo)+ ''
        skuP =  Paragraph("<para fontSize=8>{0}</para>".format(sku),styles['Normal'])
        pd= Paragraph("<para fontSize=10><b>{0}</b></para>".format(name),styles['Normal'])
        tableData.append([pd,item.priceDuringOrder,item.qty,item.discountAmount,(item.priceDuringOrder-item.discountAmount) * item.qty])
        tableDataPackingSlip.append([sku,pd,item.qty])

    if not packingSlip:
        tableData.append(['TOTAL','','','',total])
        t1=Table(tableData,colWidths=[1.7*inch , 0.5*inch , 0.5*inch, 0.7*inch , 0.7*inch])
    else:
        t1=Table(tableDataPackingSlip,colWidths=[2*inch , 3.2*inch , 0.5*inch])

    # t1=Table(tableData,colWidths=[1.7*inch , 0.5*inch , 0.5*inch, 0.7*inch , 0.7*inch])
    t1.setStyle(TableStyle([('FONTSIZE', (0, 0), (-1, -1), 8),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black),('BOX', (0,0), (-1,-1), 0.25, colors.black),('VALIGN',(0,0),(-1,-1),'TOP'), ]))
    elements.append(t1)
    elements.append(Spacer(1, 10))

    if not packingSlip:
        if order.paymentMode != 'card':
            txt7 = '<para size=15 leftIndent=150 rightIndent=150><b>CASH TO BE COLLECT &nbsp; {0} INR</b></para>'.format(total)
            elements.append(Paragraph(txt7, styles['Normal']))
        elements.append(Spacer(1, 20))

    txt8 = '<para size=10 leftIndent=150 rightIndent=150><b>Tracking ID. : </b>{0} </para>'.format(courierAWBNo)
    elements.append(Paragraph(txt8, styles['Normal']))
    elements.append(Spacer(1, 10))
    barVal = str(courierAWBNo)
    # print barVal,'bar vallllllllllllllllllllllll'
    barcode=code39.Extended39(barVal,barWidth=0.4*mm,barHeight=10*mm,checksum=0)
    elements.append(Indenter(left=140))
    elements.append(barcode)
    elements.append(Indenter(left=-140))
    elements.append(Spacer(1, 10))
    txt9 = '<para size=10 leftIndent=150 rightIndent=150><b>Order ID. : </b>{0} </para>'.format(invNo)
    elements.append(Paragraph(txt9, styles['Normal']))
    elements.append(Spacer(1, 5))


    doc.build(elements)

def generateBillData(orderData):
    toSend = {}
    try:
        companyName = appSettingsField.objects.get(app__id=25,name='companyName').value
        companyAddress = appSettingsField.objects.get(app__id=25,name='companyAddress').value
        soup1 = BeautifulSoup(companyName)
        companyName = str(soup1.text)
        soup2 = BeautifulSoup(companyAddress)
        companyAddress = str(soup2.text)
    except:
        companyName = ''
        companyAddress = ''
    try:
        toSend['companyName'] = companyName
        toSend['companyAddress'] = companyAddress
        toSend['storeName'] = ''
        toSend['storeAddress'] = ''
        toSend['storeAddress'] = ''
        toSend['products'] = []
        toSend['doubleCopyData'] = []
        grandTotal = 0
        modeOfPayment = ''
        barCVal = ''
        for item in orderData.orderQtyMap.all():
            if item.status != 'cancelled':
                grandTotal += item.totalAmount
                modeOfPayment = item.modeOfPayment
                barCVal = item.courierAWBNo

                if item.prodSku != item.product.product.serialNo:
                    product = ProductVerient.objects.get(sku=item.prodSku)
                    qtyData = product.unitPerpack * item.product.product.howMuch
                else:
                    qtyData =  item.product.product.howMuch

                if str(item.product.product.unit)=='Gram' or str(item.product.product.unit)=='gm':
                    if qtyData >1000:
                        qtyValue = str(qtyData/1000) + ' Kg'
                    else:
                        qtyValue = str(qtyData) + ' gm'
                elif str(item.product.product.unit)=='Millilitre' or str(item.product.product.unit)=='ml':
                    if qtyData>1000:
                        qtyValue = str(qtyData/1000) + ' lt'
                    else:
                        qtyValue = str(qtyData) + ' ml'
                elif str(item.product.product.unit)=='Size and Color' or str(item.product.product.unit)=='Size':
                    if int(qtyData)==1:
                        qtyValue = 'XS'
                    elif int(qtyData)==2:
                        qtyValue = 'S'
                    elif int(qtyData)==2:
                        qtyValue = 'M'
                    elif int(qtyData)==4:
                        qtyValue = 'L'
                    elif int(qtyData)==5:
                        qtyValue = 'XL'
                    elif int(qtyData)==6:
                        qtyValue = 'XL'
                    else:
                        qtyValue = qtyData
                else:
                  qtyValue = qtyData

                if item.desc:
                     desc = item.desc
                else:
                    desc =""

                name = str(item.product.product.name)+ ' ' + str(qtyValue)+ ' ' +str(desc)

                if item.prodSku is not None:
                    if item.prodSku == item.product.product.serialNo:
                        price = item.product.product.price
                    else:
                        prod = ProductVerient.objects.filter(sku = item.prodSku)
                        if len(prod)>0:
                            price = prod[0].price
                        else:
                            price = item.product.product.price
                else:
                    price = item.product.product.price

                if item.product.product.productMeta:
                    pm = item.product.product.productMeta
                    productMeta = {'typ':pm.typ,'code':pm.code,'taxRate':pm.taxRate}
                else:
                    productMeta = None
                qty = item.qty
                toSend['products'].append({"data":{"product":{"productMeta":productMeta,"price":price},"name":name},"quantity":qty})
                toSend['doubleCopyData'].append({"name":name,"qty":qty})

        toSend['barCVal'] = barCVal
        toSend['modeOfPayment'] = modeOfPayment
        toSend['grandTotal'] = grandTotal
        toSend['amountRecieved'] = grandTotal
        toSend['products'] = json.dumps(toSend['products'])
    except:
        print 'errorrrrrrrrrrr'
    return toSend

class DownloadManifestAPI(APIView):
    renderer_classes = (JSONRenderer,)
    # permission_classes = (permissions.IsAuthenticated ,)
    def get(self , request , format = None):
        # print self.request.GET
        if 'trackingId' in request.GET:
            try:
                qtyObjs = OrderQtyMap.objects.filter(courierAWBNo=request.GET['trackingId'])
                for i in qtyObjs:
                    # print i.status,'statusssssssssss'
                    if i.status == 'created':
                        i.status = 'packed'
                        i.save()
            except:
                pass
            fd = open(os.path.join(globalSettings.BASE_DIR, 'media_root/ecommerce/manifest','example_shipment_label'+request.GET['trackingId']+'.pdf'))
            response = HttpResponse(fd, content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename=manifest.pdf'
            return response
        else:
            if 'allData' in request.GET:
                item = Order.objects.get(pk=request.GET['allData'])
                typ='all'
                barCVal = ''
                for i in item.orderQtyMap.all():
                    if i.courierAWBNo:
                        barCVal = i.courierAWBNo
            elif 'qPk' in request.GET:
                item = OrderQtyMap.objects.get(pk = request.GET['qPk'])
                typ='one'
                barCVal = item.courierAWBNo
            else:
                return Response({}, status = status.HTTP_200_OK)

            ab = appSettingsField.objects.filter(name='posPrinting')
            if 'packingSlip' in request.GET:
                packingSlip = True
            else:
                packingSlip = False

            if 'printerDeviceId' in request.GET and len(ab)>0:
                if ab[0].flag:
                    print 'valid printerrrrrrrrrrrr idddddddddd'
                    resData = generateBillData(item)
                    resData['id'] = item.pk
                    requests.post("http://"+globalSettings.WAMP_SERVER+":8080/notify",json={'topic': 'service.POS.Printer.{0}'.format(request.GET['printerDeviceId']),'args': [{'data':resData}]})
                    try:
                        doubleCopyObj = appSettingsField.objects.get(name='doubleInvoiceCopy')
                        if doubleCopyObj.flag:
                            print 'small slipppppppppp'
                            doubleCopyData = resData['doubleCopyData']
                            for i in doubleCopyData:
                                # print 'product data',i
                                requests.post("http://"+globalSettings.WAMP_SERVER+":8080/notify",json={'topic': 'service.POS.Printer.{0}'.format(request.GET['printerDeviceId']),'args': [{'id':item.pk,'name':i['name'],'qty':i['qty'],'smallBill':'yes'}]})
                    except:
                        pass


                    return Response({},status=status.HTTP_200_OK)

            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = 'attachment;filename="manifest.pdf"'
            manifest(response,item,typ,packingSlip)
            return response

class PostBarcodeAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def post(self , request , format = None):
        print request.data
        if 'barcode' in request.data and 'userId' in request.data and 'typ' in request.data:
            if request.data['typ'] == 'online':
                qtyObjs = OrderQtyMap.objects.filter(courierAWBNo=request.data['barcode'])
                print qtyObjs.count()
                if qtyObjs.count()>0:
                    try:
                        userObj = User.objects.get(pk=int(request.data['userId']))
                        for i in qtyObjs:
                            i.status = 'delivered'
                            i.orderBy = userObj
                            if 'modeOfPayment' in request.data:
                                i.modeOfPayment = request.data['modeOfPayment']
                                i.paidAmount = i.totalAmount
                            i.save()
                    except:
                        return Response({'statusText':'Invalid UserId'}, status = status.HTTP_200_OK)
                    parentObj = qtyObjs[0].order.get()
                    childObjs = parentObj.orderQtyMap.all()
                    print childObjs.count(),'child countttttt'
                    for i in childObjs:
                        if i.status != 'delivered':
                            print 'not all products are delivered yet'
                            break
                    else:
                        print 'all products are delivered  so changing order status to completed'
                        parentObj.status = 'completed'
                        parentObj.save()
                else:
                    return Response({'statusText':'Invalid Barcode'}, status = status.HTTP_200_OK)
            elif request.data['typ'] == 'offline':
                try:
                    invObj = Invoice.objects.get(barCVal=request.data['barcode'])
                    try:
                        userObj = User.objects.get(pk=int(request.data['userId']))
                        invObj.orderBy = userObj
                        invObj.status = 'Completed'
                        if 'modeOfPayment' in request.data:
                            invObj.modeOfPayment = request.data['modeOfPayment']
                            invObj.amountRecieved = invObj.grandTotal
                        invObj.save()
                    except:
                        return Response({'statusText':'Invalid UserId'}, status = status.HTTP_200_OK)
                except:
                    return Response({'statusText':'Invalid UserId'}, status = status.HTTP_200_OK)
            else:
                return Response({'statusText':'Invalid Data'}, status = status.HTTP_200_OK)
        else:
            return Response({'statusText':'Invalid Data'}, status = status.HTTP_200_OK)
        return Response({'statusText':'Status Updated'}, status = status.HTTP_200_OK)

class GetDeliveredOrderDetailsAPI(APIView):
    renderer_classes = (JSONRenderer,)
    def get(self, request, format=None):
        print request.GET
        if 'userId' in request.GET and 'dated' in request.GET:
            toRet = {'statusText':'Success'}
            dts = request.GET['dated']
            # dt = datetime.datetime.strptime(dts,'%Y-%m-%d').date()
            userObj = User.objects.get(pk=int(request.GET['userId']))
            qtyObjs = OrderQtyMap.objects.filter(orderBy=userObj,updated__icontains=dts)
            invObjs = Invoice.objects.filter(orderBy=userObj,updated__icontains=dts)

            deliveredCount = qtyObjs.filter(status='delivered').count() + invObjs.filter(status='Completed').count()
            ongoingCount = qtyObjs.filter(status='outForDelivery').count() + invObjs.filter(status='outForDelivery').count()
            toRet['deliveredCount'] = deliveredCount
            toRet['ongoingCount'] = ongoingCount

            qtyMCodAmount = qtyObjs.filter(modeOfPayment='cash').aggregate(tot=Sum('paidAmount'))['tot']
            qtyMCodAmount = qtyMCodAmount if qtyMCodAmount else 0
            qtyMCardAmount = qtyObjs.filter(modeOfPayment='card').aggregate(tot=Sum('paidAmount'))['tot']
            qtyMCardAmount = qtyMCardAmount if qtyMCardAmount else 0

            invMCodAmount = invObjs.filter(modeOfPayment='cash').aggregate(tot=Sum('amountRecieved'))['tot']
            invMCodAmount = invMCodAmount if invMCodAmount else 0
            invMCardAmount = invObjs.filter(modeOfPayment='card').aggregate(tot=Sum('amountRecieved'))['tot']
            invMCardAmount = invMCardAmount if invMCardAmount else 0

            cod = qtyMCodAmount + invMCodAmount
            card = qtyMCardAmount + invMCardAmount
            total = cod + card

            toRet['cod'] = cod
            toRet['card'] = card
            toRet['total'] = total
        else:
            return Response({'statusText':'Invalid Data'}, status = status.HTTP_200_OK)

        return Response(toRet,status = status.HTTP_200_OK)

class SendStatusAPI(APIView):
    renderer_classes = (JSONRenderer,)
    def post(self , request , format = None):
        emailAddr=[]
        print request.data['value'],'aaaaaaaaaaaaaa'
        oq = OrderQtyMap.objects.filter(pk = request.data['value'])
        for i in oq:
            productId = i.pk
            productStatus = i.status
            productName = i.product.product.name
            qty = i.qty
        print str(productName)
        o = Order.objects.filter(orderQtyMap = productId)
        for j in o:
            orderId = j.pk
            emailAddr.append(j.user.email)
        # emailAddr.append(customer.email)
        email_subject = "Order Details:"
        if productStatus == 'created':
            msgBody = "Your order with  product name: " + str(productName) + ", Quantity of: " +str(qty)+" has been placed"
        elif productStatus == 'packed':
            msgBody = "Your order status with  product name: " + str(productName) + ", Quantity of: " +str(qty)+" has been packed"
        elif productStatus == 'shipped':
            msgBody = "Your order status with  product name: " + str(productName) + ", Quantity of: " +str(qty)+" has been changed from packed to shipped"
        elif productStatus == 'inTransit':
            msgBody = "Your order status with  product name: " + str(productName) + ", Quantity of: " +str(qty)+" has been changed from shipped to In Transit"
        elif productStatus == 'reachedNearestHub':
            msgBody = "Your order with  product name: " + str(productName) + ", Quantity of: " +str(qty)+" has been Reached to the nearest Hub"
        elif productStatus == 'outForDelivery':
            msgBody = "Your order with  product name: " + str(productName) + ", Quantity of: " +str(qty)+" is Out for delivery"
        elif productStatus == 'cancelled':
            msgBody = "Your order with  product name: " + str(productName) + ", Quantity of: " +str(qty)+" has been cancelled"
        elif productStatus == 'returnToOrigin':
            msgBody = "Product has been returned to origin"
        elif productStatus == 'returned':
            msgBody ="Product has been returned"

        print emailAddr[0],'ffffffffffff'
        email_body = msgBody
        email_subject = "Order Status"
        email_to = [emailAddr[0]]
        # for i in globalSettings.G_ADMIN:
        #     email_to.append(str(i))
        email_cc = []
        email_bcc = []
        send_email(email_body,email_to,email_subject,email_cc,email_bcc,'text')

        return Response({}, status = status.HTTP_200_OK)

class SendDeliveredStatus(APIView):
    renderer_classes = (JSONRenderer,)
    def post(self , request , format = None):
        emailAddr=[]
        oq = []
        o = []
        m=[]
        promoAmount=0
        oq=OrderQtyMap.objects.get(pk = request.data['value'])
        print oq.pk
        price = oq.product.product.price - (oq.product.product.discount * oq.product.product.price)/100
        total = oq.qty * price
        o=Order.objects.get(orderQtyMap = oq.pk)
        emailAddr.append(o.user.email)
        promoObj = Promocode.objects.all()
        for p in promoObj:
            if str(p.name)==str(o.promoCode):
                promoAmount = p.discount
        print promoAmount
        grandTotal=total-(promoAmount * total)/100
        grandTotal=round(grandTotal, 2)
        attachment =  oq.product.files.values_list('attachment', flat=True)
        print '**************************'
        try:
            companyAddress = appSettingsField.objects.filter(name='companyAddress')[0].value
            cleanr = re.compile('<.*?>')
            companyAddress = re.sub(cleanr, '', companyAddress)
        except:
            companyAddress = ''
        try:
            gstValue = appSettingsField.objects.filter(name='cstNo')[0].value
        except:
            gstValue = ''
        ctx = {
            'heading' : "Invoice Details",
            'linkUrl': globalSettings.BRAND_NAME,
            'sendersAddress' : globalSettings.SEO_TITLE,
            'promoAmount':promoAmount,
            'attachment':"https://media/"+attachment[0],
            'grandTotal':grandTotal,
            'total':total,
            'order': o,
            'price':price,
            'orderQTY':oq,
            'linkedinUrl' : lkLink,
            'fbUrl' : fbLink,
            'twitterUrl' : twtLink,
            'companyAddress':companyAddress,
            'gstValue':gstValue
        }
        email_body = get_template('app.ecommerce.deliveryDetailEmail.html').render(ctx)
        email_subject = "Delivered"
        email_to = []
        email_to.append(str(o.user.email))
        email_cc = []
        email_bcc = []
        send_email(email_body,email_to,email_subject,email_cc,email_bcc,'html')
        return Response({}, status = status.HTTP_200_OK)


class SendFeedBackAPI(APIView):
    renderer_classes = (JSONRenderer,)
    def post(self , request , format = None):
        emailAddr=[]
        response=''
        supportObj=SupportFeed.objects.get(pk = request.data['datapk'])
        response = request.data['response']
        # responseData = BeautifulSoup(response, 'html.parser')
        responseData = Markup(response)
        emailAddr.append(str(supportObj.email))
        ctx = {
            'heading' : "On response to your Feed Back",
            'linkUrl': globalSettings.BRAND_NAME,
            'sendersAddress' : globalSettings.SEO_TITLE,
            'question' : supportObj.message,
            'response':responseData,
            'linkedinUrl' : lkLink,
            'fbUrl' : fbLink,
            'twitterUrl' : twtLink,
        }
        print ctx
        email_body = get_template('app.ecommerce.support.email.html').render(ctx)
        email_subject = "Response"
        email_cc = []
        email_bcc = []
        send_email(email_body,emailAddr,email_subject,email_cc,email_bcc,'html')
        return Response({}, status = status.HTTP_200_OK)


themeColor = colors.HexColor(globalSettings.ECOMMERCE_THEME)

styles=getSampleStyleSheet()
styleN = styles['Normal']
styleH = styles['Heading1']

try:
    settingsFields = application.objects.get(name = 'app.public.ecommerce').settings.all()
except:
    print "ERROR : settingsFields = application.objects.get(name = 'app.public.ecommerce').settings.all()"

try:
    ecommerceSetting = application.objects.get(name = 'app.public.ecommerce').settings.all()
except:
    print "ERROR : application.objects.get(name = 'app.public.ecommerce').settings.get(name__iexact = 'gstEnabled')"

class FullPageImage(Flowable):
    def __init__(self , img):
        Flowable.__init__(self)
        self.image = img

    def draw(self):
        img = utils.ImageReader(self.image)

        iw, ih = img.getSize()
        aspect = ih / float(iw)
        width, self.height = PAGE_SIZE
        width -= 3.5*cm
        self.canv.drawImage(os.path.join(BASE_DIR , self.image) , -1 *MARGIN_SIZE + 1.5*cm , -1* self.height + 5*cm , width, aspect*width)

class expanseReportHead(Flowable):

    def __init__(self, request , contract):
        Flowable.__init__(self)
        self.req = request
        self.contract = contract
    #----------------------------------------------------------------------
    def draw(self):
        """
        draw the floable
        """
        now = datetime.datetime.now(pytz.timezone('Asia/Kolkata'))
        docTitle = 'ORDER INVOICE'
        dateCreated=self.contract.created.strftime("%d-%m-%Y")
        passKey = '%s%s'%(str(self.req.user.date_joined.year) , self.req.user.pk) # also the user ID
        docID = '%s%s' %( now.year , self.contract.pk)


        pSrc = '''
        <font size=10>&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp&nbsp<u>%s</u></font><br/><br/><br/>
        <font size=8>
        <strong>Dated:</strong> %s<br/><br/>
        <strong>Invoice ID:</strong> %s<br/><br/>
        </font>
        ''' % ( docTitle  , dateCreated , docID)

        story = []
        head = Paragraph(pSrc , styleN)
        head.wrapOn(self.canv , 200*mm, 50*mm)
        head.drawOn(self.canv , 0*mm, -10*mm)

        # barcode_value = "1234567890"
        # barcode39 = barcode.createBarcodeDrawing('EAN13', value = barcode_value,barWidth=0.3*mm,barHeight=10*mm)
        #
        # barcode39.drawOn(self.canv,160*mm,0*mm)
        # self.canv.drawImage(os.path.join(BASE_DIR , 'logo.png') , 80*mm , 0*mm , 2*cm, 2*cm)

def addPageNumber(canvas, doc):
    """
    Add the page number
    """
    print doc.contract
    now = datetime.datetime.now(pytz.timezone('Asia/Kolkata'))
    passKey = '%s%s'%(str(doc.request.user.date_joined.year) , doc.request.user.pk) # also the user ID
    docID = '%s%s' %(now.year , doc.contract.pk)

    qrw = QrCodeWidget('http://cioc.co.in/documents?id=%s&passkey=%s&app=crmInvoice' %(docID , passKey))
    b = qrw.getBounds()

    w=b[2]-b[0]
    h=b[3]-b[1]

    d = Drawing(60,60,transform=[60./w,0,0,60./h,0,0])
    # d.add(qrw)
    renderPDF.draw(d, canvas ,180*mm,270*mm)

    pass

    # page_num = canvas.getPageNumber()
    # text = "<font size='8'>Page #%s</font>" % page_num
    # p = Paragraph(text , styleN)
    # p.wrapOn(canvas , 50*mm , 10*mm)
    # p.drawOn(canvas , 100*mm , 10*mm)



class PageNumCanvas(canvas.Canvas):

    #----------------------------------------------------------------------
    def __init__(self, *args, **kwargs):
        """Constructor"""
        canvas.Canvas.__init__(self, *args, **kwargs)
        self.pages = []

    #----------------------------------------------------------------------
    def showPage(self):
        """
        On a page break, add information to the list
        """
        self.pages.append(dict(self.__dict__))
        self._startPage()

    #----------------------------------------------------------------------
    def save(self):
        """
        Add the page number to each page (page x of y)
        """
        page_count = len(self.pages)

        for page in self.pages:
            self.__dict__.update(page)
            # self.draw_page_number(page_count)
            self.drawLetterHeadFooter()
            canvas.Canvas.showPage(self)

        canvas.Canvas.save(self)


    #----------------------------------------------------------------------
    def draw_page_number(self, page_count):
        """
        Add the page number
        """

        text = "<font size='8'>Page #%s of %s</font>" % (self._pageNumber , page_count)
        p = Paragraph(text , styleN)
        p.wrapOn(self , 50*mm , 10*mm)
        p.drawOn(self , 100*mm , 10*mm)

    def drawLetterHeadFooter(self):
        self.setStrokeColor(themeColor)
        self.setFillColor(themeColor)
        self.rect(0,0,1500,70, fill=True)
        # print dir(self)
        compNameStyle = styleN.clone('footerCompanyName')
        compNameStyle.textColor = colors.white;

        p = Paragraph(settingsFields.get(name = 'companyName').value , compNameStyle)
        # p = Paragraph('Business Network International' , compNameStyle)

        p.wrapOn(self , 50*mm , 10*mm)
        p.drawOn(self , 85*mm  , 18*mm)

        p1 = Paragraph(settingsFields.get(name = 'companyAddress').value , compNameStyle)
        p1.wrapOn(self , 150*mm , 10*mm)
        p1.drawOn(self , 55*mm  , 10*mm)


        p2 = Paragraph( settingsFields.get(name = 'contactDetails').value, compNameStyle)
        p2.wrapOn(self , 200*mm , 10*mm)
        p2.drawOn(self , 40*mm  , 1*mm)

        ab = appSettingsField.objects.filter(name='cstNo')
        try:
            gstin = ab[0].value
            ab = appSettingsField.objects.filter(name='isStoreGlobal')
            if len(ab)>0:
                if ab[0].flag:
                    gstin = ''
        except :
            gstin = ''
        p4 = Paragraph(gstin , compNameStyle)
        p4.wrapOn(self , 200*mm , 10*mm)
        p4.drawOn(self , 85*mm  , 5*mm)

        brandLogo = globalSettings.INVOICE_LOGO.split('static/')[1]
        print os.path.join(globalSettings.BASE_DIR , 'static_shared',brandLogo)
        drawing = svg2rlg(os.path.join(globalSettings.BASE_DIR , 'static_shared', brandLogo))
        sx=sy=0.5
        drawing.width,drawing.height = drawing.minWidth()*sx, drawing.height*sy
        drawing.scale(sx,sy)
        #if you want to see the box around the image
        # drawing._showBoundary = True
        renderPDF.draw(drawing, self,10*mm  , self._pagesize[1]-30*mm)

        #width = self._pagesize[0]
        # page = "Page %s of %s" % (, page_count)
        # self.setFont("Helvetica", 9)
        # self.drawRightString(195*mm, 272*mm, page)

def genInvoice(response, contract, request):
    MARGIN_SIZE = 8 * mm
    PAGE_SIZE = A4

    # c = canvas.Canvas("hello.pdf")
    # c.drawString(9*cm, 19*cm, "Hello World!")

    pdf_doc = SimpleDocTemplate(response, pagesize=PAGE_SIZE,
                                leftMargin=MARGIN_SIZE, rightMargin=MARGIN_SIZE,
                                topMargin=4 * MARGIN_SIZE, bottomMargin=3 * MARGIN_SIZE)

    # data = [['', '', '', 'Grand Total', '' , pFooterGrandTotal]]
    pdf_doc.contract = contract
    pdf_doc.request = request


    tableHeaderStyle = styles['Normal'].clone('tableHeaderStyle')
    tableHeaderStyle.textColor = colors.white
    tableHeaderStyle.fontSize = 7

# ecommerceSetting.get(name = 'gstEnabled').flag
    # totalQuant = 0
    # totalTax = 0


    if contract.billingState == contract.state:
        interState = True
    else:
        interState = False
    totaldiscount = 0
    total = 0
    totalprice = 0
    promoAmount = 0
    discount = 0
    promoCode = ''
    tableBodyStyle = styles['Normal'].clone('tableBodyStyle')
    tableBodyStyle.fontSize = 7
    tableGrandStyle = tableHeaderStyle.clone('tableGrandStyle')
    tableGrandStyle.fontSize = 10
    promoObj = Promocode.objects.all()
    if contract.promoCode:
        for p in promoObj:
            if str(p.name)==str(contract.promoCode):
                promoAmount = p.discount
                print  p.name,'aaaaaaaaaaaaaaaa'
                promoCode = p.name
    else:
        promoCode="None"
    tableData=[['Product','Quantity','Price','Total Price']]
    isStoreGlobal = False
    ab = appSettingsField.objects.filter(name='isStoreGlobal')
    currencySymbol = appSettingsField.objects.filter(name='currencySymbol')
    if len(currencySymbol)>0:
        if currencySymbol[0].value == 'fa-usd':
            currency = '(USD)'
        else:
            currency = '(INR)'
    else:
        currency = '(USD)'
    if len(ab)>0:
        if ab[0].flag:
            isStoreGlobal = True
    if not isStoreGlobal:
        if interState:
            tableData=[['Product','HSN','Quantity','Price' + currency,'SGST (%)','CGST (%)','Total GST'+currency,'Total Price'+currency]] #here
        else:
            tableData=[['Product','HSN','Quantity','Price' + currency,'IGST (%)','IGST'+currency,'Total Price'+currency]] #here
    for i in contract.orderQtyMap.all():
        print i.desc,'ssssssssssssss'
        if str(i.status)!='cancelled':
            if i.prodSku == i.product.product.serialNo:
                print i.product.product.name, i.product.product.discount, i.priceDuringOrder
                price = i.priceDuringOrder - (i.product.product.discount * i.priceDuringOrder)/100
                price=round(price, 2)
                totalprice = i.qty*price
                totalprice=round(totalprice, 2)
                total=round(total, 2)
                if i.desc:
                    desc =i.desc
                else:
                    desc=""
                qtyData = i.product.product.howMuch
                if str(i.product.product.unit)=='Gram' or str(i.product.product.unit)=='gm':
                    if qtyData >1000:
                        qtyValue = str(qtyData/1000) + ' Kg'
                    else:
                        qtyValue = str(qtyData) + ' gm'
                elif str(i.product.product.unit)=='Millilitre' or str(i.product.product.unit)=='ml':
                    if qtyData>1000:
                        qtyValue = str(qtyData/1000) + ' lt'
                    else:
                        qtyValue = str(qtyData) + ' ml'
                elif str(i.product.product.unit)=='Size and Color' or str(i.product.product.unit)=='Size':
                    if int(qtyData)==1:
                        qtyValue = 'XS'
                    elif int(qtyData)==2:
                        qtyValue = 'S'
                    elif int(qtyData)==2:
                        qtyValue = 'M'
                    elif int(qtyData)==4:
                        qtyValue = 'L'
                    elif int(qtyData)==5:
                        qtyValue = 'XL'
                    elif int(qtyData)==6:
                        qtyValue = 'XL'
                    else:
                        qtyValue = qtyData
                else:
                  qtyValue = qtyData
                name = str(i.product.product.name) + ' '  + str(qtyValue) + ' ' +str(desc)
                if not isStoreGlobal:
                    if i.product.product.productMeta is None:
                        if interState:
                            tableData.append([name,'',i.qty,price,'','','',totalprice]) #herre
                        else:
                            tableData.append([name,'',i.qty,price,'','',totalprice]) #herre
                    else:
                        totalprice = totalprice + (i.product.product.productMeta.taxRate * totalprice)/100
                        gstval = (i.product.product.productMeta.taxRate * price )/100
                        if interState:
                            tableData.append([name,i.product.product.productMeta.code,i.qty,price,i.product.product.productMeta.taxRate/2,i.product.product.productMeta.taxRate/2,gstval,totalprice])
                        else:
                            tableData.append([name,i.product.product.productMeta.code,i.qty,price,i.product.product.productMeta.taxRate,gstval,totalprice])

                else:
                    tableData.append([name,i.qty,price,totalprice])
                total+=totalprice
            else:
                prodData = ProductVerient.objects.get(sku = i.prodSku)
                price = prodData.discountedPrice
                price=round(price, 2)
                totalprice = i.qty*price
                totalprice=round(totalprice, 2)
                total=round(total, 2)
                if i.desc:
                    desc =i.desc
                else:
                    desc=""
                qtyData = i.product.product.howMuch * prodData.unitPerpack
                if str(i.product.product.unit)=='Gram' or str(i.product.product.unit)=='gm':
                    if qtyData >1000:
                        qtyValue = str(qtyData/1000) + ' Kg'
                    else:
                        qtyValue = str(qtyData) + ' gm'
                elif str(i.product.product.unit)=='Millilitre' or str(i.product.product.unit)=='ml':
                    if qtyData>1000:
                        qtyValue = str(qtyData/1000) + ' lt'
                    else:
                        qtyValue = str(qtyData) + ' ml'
                elif str(i.product.product.unit)=='Size and Color' or str(i.product.product.unit)=='Size':
                    if int(qtyData)==1:
                        qtyValue = 'XS'
                    elif int(qtyData)==2:
                        qtyValue = 'S'
                    elif int(qtyData)==2:
                        qtyValue = 'M'
                    elif int(qtyData)==4:
                        qtyValue = 'L'
                    elif int(qtyData)==5:
                        qtyValue = 'XL'
                    elif int(qtyData)==6:
                        qtyValue = 'XL'
                    else:
                        qtyValue = qtyData
                else:
                  qtyValue = qtyData


                name = str(i.product.product.name) + ' ' + str(qtyValue)+ ' ' +str(desc)

                if i.product.product.productMeta is None:
                    if interState:
                        tableData.append([name,'',i.qty,price,'','','',totalprice]) #herre
                    else:
                        tableData.append([name,'',i.qty,price,'','',totalprice]) #herre
                else:
                    totalprice = totalprice + (i.product.product.productMeta.taxRate * totalprice)/100
                    gstval = (i.product.product.productMeta.taxRate * price )/100
                    if interState:
                        tableData.append([name,i.product.product.productMeta.code,i.qty,price,i.product.product.productMeta.taxRate/2,i.product.product.productMeta.taxRate/2,gstval,totalprice])
                    else:
                        tableData.append([name,i.product.product.productMeta.code,i.qty,price,i.product.product.productMeta.taxRate,gstval,totalprice])
                total+=totalprice
    shippingCharges = contract.shippingCharges
    grandTotal=total-(promoAmount * total)/100
    grandTotal=round(grandTotal + shippingCharges, 2)
    if not isStoreGlobal:
        tableData.append(['','','','TOTAL' + currency,'','',total])
        tableData.append(['','','','COUPON APPLIED(%)','',promoCode,promoAmount])
        tableData.append(['','','','SHIPPING CHARGES' + currency,'','',shippingCharges])
        tableData.append(['','','','GRAND TOTAL'+ currency,'','',grandTotal])
        if interState:
            t1=Table(tableData,colWidths=[3.2*inch , 0.5*inch , 0.6*inch, 0.7*inch,0.6*inch,0.6*inch,1*inch , 1*inch]) #herre
        else:
            t1=Table(tableData,colWidths=[3.2*inch , 0.6*inch , 0.7*inch, 0.7*inch,0.7*inch,1*inch , 1*inch]) #herre
    else:
        tableData.append(['','TOTAL'+currency,'',total])
        tableData.append(['','COUPON APPLIED(%)',promoCode,promoAmount])
        tableData.append(['','SHIPPING CHARGES','',shippingCharges])
        tableData.append(['','','GRAND TOTAL'+currency,grandTotal])
        t1=Table(tableData,colWidths=[3.2*inch , 1.5*inch , 1.5*inch, 0.8*inch , 1.5*inch])

    t1.setStyle(TableStyle([('FONTSIZE', (0, 0), (-1, -1), 8),('INNERGRID', (0,0), (-1,-1), 0.25,  colors.HexColor('#bdd3f4')),('INNERGRID', (0,-4), (-1,-1), 0.25, colors.white),('LINEABOVE', (0,-3), (-1,-1), 0.25, colors.HexColor('#bdd3f4')),('INNERGRID', (0,-1), (-1,-1), 0.25, colors.white),('INNERGRID', (0,-2), (-1,-1), 0.25, colors.white),('INNERGRID', (0,-1), (-1,-1), 0.25, colors.white),('LINEABOVE', (0,-1), (-1,-1), 0.25, colors.black),('INNERGRID', (0,-3), (-1,-1), 0.25, colors.white),('LINEABOVE', (0,-2), (-1,-1), 0.25, colors.HexColor('#bdd3f4')),('BOX', (0,0), (-1,-1), 0.25,  colors.HexColor('#bdd3f4')),('VALIGN',(0,0),(-1,-1),'TOP'),('BACKGROUND', (0, 0), (-1, 0),colors.HexColor('#f0f0f0')) ]))
    if ecommerceSetting.get(name = 'gstEnabled').flag == True:
        gst = """
        <font size='6'><strong>GST : </strong></font>
        """
        try:
            detailsObj = contract.user.profile.details
            details = ast.literal_eval(detailsObj)
            gstVal = details['GST']
        except:
            gstVal = ''
    else:
        gst = ''
        gstVal = ''

    story = []

    expHead = expanseReportHead(request, contract)
    story.append(Spacer(2.5, 2 * cm))
    story.append(expHead)
    story.append(Spacer(2.5, 0.75 * cm))

    summryParaSrc3 = """
    <font size='8'><strong>Customer details:</strong></font> <br/>
    """
    story.append(Paragraph(summryParaSrc3 , styleN))

    if isStoreGlobal:
        summryParaSrc = Paragraph("""
        <para backColor = '#ffffff' leftIndent = 10>
        <font size='6'><strong>Your Billing Address:</strong></font> <br/>
        <font size='6'>
        %s %s<br/>
        %s <br/>
        %s <br/>
        %s %s - %s<br/>
        %s <br/>
        %s
        </font>
        </para>
        """ %(contract.user.first_name , contract.user.last_name , contract.billingLandMark , contract.billingStreet , contract.billingCity , contract.billingState , contract.billingPincode, contract.country, contract.mobileNo),styles['Normal'])
    else:
        summryParaSrc = Paragraph("""
        <para backColor = '#ffffff' leftIndent = 10>
        <font size='6'><strong>Your Billing Address:</strong></font> <br/>
        <font size='6'>
        %s %s<br/>
        %s <br/>
        %s <br/>
        %s %s - %s<br/>
        %s <br/>
        %s<br/>
        %s  %s
        </font>
        </para>
        """ %(contract.user.first_name , contract.user.last_name , contract.billingLandMark , contract.billingStreet , contract.billingCity , contract.billingState , contract.billingPincode, contract.country, contract.mobileNo,gst,gstVal),styles['Normal'])


    if isStoreGlobal:
        summryParaSrc1 = Paragraph("""
        <para backColor = #ffffff leftIndent = 10>
        <font size='6'><strong>Your Shipping Address:</strong></font> <br/>
        <font size='6'>
        %s %s<br/>
        %s ,
        %s <br/>
        %s %s - %s<br/>
        %s<br/>
        %s
        </font></para>
        """ %(contract.user.first_name , contract.user.last_name , contract.landMark , contract.street , contract.city , contract.state , contract.pincode, contract.country, contract.mobileNo),styles['Normal'])
    else:
        summryParaSrc1 = Paragraph("""
        <para backColor = #ffffff leftIndent = 10>
        <font size='6'><strong>Your Shipping Address:</strong></font> <br/>
        <font size='6'>
        %s %s<br/>
        %s ,
        %s <br/>
        %s %s - %s<br/>
        %s<br/>
        %s<br/>
        %s %s
        </font></para>
        """ %(contract.user.first_name , contract.user.last_name , contract.landMark , contract.street , contract.city , contract.state , contract.pincode, contract.country, contract.mobileNo,gst,gstVal),styles['Normal'])




    td=[[summryParaSrc,' ',summryParaSrc1]]
    # story.append(Paragraph(summryParaSrc , styleN))
    t=Table(td,colWidths=[3*inch , 1*inch , 3*inch])
    t.setStyle(TableStyle([('BACKGROUND', (0, 0), (0, 0),colors.HexColor('#ffffff')),('BACKGROUND', (-1, -1), (-1,-1 ),colors.HexColor('#ffffff')) ]))
    story.append(t)
    story.append(Spacer(2.5,0.5*cm))

    orderNumber = contract.pk

    summryParaSrc13 = """
    <font size='8'><strong>Order Number: %s </strong></font> <br/> <br/>
    """ % ( orderNumber)
    story.append(Paragraph(summryParaSrc13 , styleN))


    summryParaSrc4 = """
    <font size='8'><strong>Order details:</strong></font> <br/>
    """
    story.append(Paragraph(summryParaSrc4 , styleN))
    story.append(t1)
    story.append(Spacer(2.5,0.5*cm))

    summryParaSrc2 = """
    <font size='4'><strong>Computer generated bill, no signature required</strong></font> <br/>
    """
    story.append(Paragraph(summryParaSrc2 , styleN))


    pdf_doc.build(story,onFirstPage=addPageNumber, onLaterPages=addPageNumber, canvasmaker=PageNumCanvas)


# def link_callback(uri, rel):
#     """
#     Convert HTML URIs to absolute system paths so xhtml2pdf can access those
#     resources
#     """
#     # use short variable names
#     sUrl = settings.STATIC_URL      # Typically /static/
#     sRoot = settings.STATIC_ROOT    # Typically /home/userX/project_static/
#     mUrl = settings.MEDIA_URL       # Typically /static/media/
#     mRoot = settings.MEDIA_ROOT     # Typically /home/userX/project_static/media/
#
#     # convert URIs to absolute system paths
#     if uri.startswith(mUrl):
#         path = os.path.join(mRoot, uri.replace(mUrl, ""))
#     elif uri.startswith(sUrl):
#         path = os.path.join(sRoot, uri.replace(sUrl, ""))
#     else:
#         return uri  # handle absolute uri (ie: http://some.tld/foo.png)
#
#     # make sure that file exists
#     if not os.path.isfile(path):
#             raise Exception(
#                 'media URI must start with %s or %s' % (sUrl, mUrl)
#             )
#     return path

# def genBNIInvoice(response,o, request):
#     template_path = 'invoice.html'
#     context = {'myvar': {'name':'vikas'}}
#     response = HttpResponse(content_type='application/pdf')
#     response['Content-Disposition'] = 'attachment; filename="report.pdf"'
#     template = get_template(template_path)
#     html = template.render(context)
#     pisaStatus = pisa.CreatePDF(
#        html, dest=response, link_callback=link_callback)
#     if pisaStatus.err:
#        return HttpResponse('We had some errors <pre>' + html + '</pre>')
#     return response



class PageCanvas(canvas.Canvas):

    #----------------------------------------------------------------------
    def __init__(self, *args, **kwargs):
        """Constructor"""
        canvas.Canvas.__init__(self, *args, **kwargs)
        self.pages = []

    #----------------------------------------------------------------------
    def showPage(self):
        """
        On a page break, add information to the list
        """
        self.pages.append(dict(self.__dict__))
        self._startPage()

    #----------------------------------------------------------------------
    def save(self):
        """
        Add the page number to each page (page x of y)
        """
        page_count = len(self.pages)

        for page in self.pages:
            self.__dict__.update(page)
            # self.draw_page_number(page_count)
            if page == self.pages[-1]:
                self.drawLetterHeadFooter()
            else:
                pass

            canvas.Canvas.showPage(self)

        canvas.Canvas.save(self)

    #----------------------------------------------------------------------
    def draw_page_number(self, page_count):
        """
        Add the page number
        """

        text = "<font size='8'>Page #%s of %s</font>" % (self._pageNumber , page_count)
        p = Paragraph(text , styleN)
        p.wrapOn(self , 50*mm , 10*mm)
        p.drawOn(self , 100*mm , 10*mm)

    def drawLetterHeadFooter(self):
        panNumber = settingsFields.get(name = 'companyAddress').value
        p = Paragraph("<para fontSize=8  alignment='center'>We Thank You for Your Business.</para>",styles['Normal'])

        p.wrapOn(self , 50*mm , 10*mm)
        p.drawOn(self , 85*mm  , 8*mm)


def generateInvoiceBni(response, contract ,request):

    isStoreGlobal = False
    ab = appSettingsField.objects.filter(name='isStoreGlobal')
    if len(ab)>0:
        if ab[0].flag:
            isStoreGlobal = True

    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(response,pagesize=letter, topMargin=0.2*cm,leftMargin=0.1*cm,rightMargin=0.1*cm)
    doc.request = request
    elements = []
    companyAddress = settingsFields.get(name = 'companyAddress').value
    companyAddress = companyAddress.replace('lutions,', 'lutions,<br/>', 1)
    companyAddress = companyAddress.replace('ross,', 'ross,<br/>', 1)
    companyAddress = companyAddress.replace('nagar,', 'nagar,<br/>', 1)
    telPhone  = str(settingsFields.get(name = 'phone').value)
    email = str(settingsFields.get(name = 'email').value)

    websiteAddress = globalSettings.SITE_ADDRESS
    # brandLogo = globalSettings.ICON_LOGO

    if isStoreGlobal:
        gstvarName = ''
        gstNumber = ''
    else:
        gstvarName = 'GSTIN is :'
        gstNumber = str(settingsFields.get(name = 'cstNo').value)


    s01 =Paragraph("<para fontSize=8> "+ companyAddress +" <br/> Tel :"+ telPhone +" <br/> Email: "+ email +" <br/> "+websiteAddress+" <br/> "+ gstvarName +" "+ gstNumber +" </para>",styles['Normal'])
    s02 =Paragraph("<para fontSize=8> </para>",styles['Normal'])
    s03 = []
    brandLogo = globalSettings.ICON_LOGO.split('static/')[1]
    logo = os.path.join(globalSettings.BASE_DIR , 'static_shared',brandLogo)
    im = GNAA(logo, width=120, height=80)
    s03.append(im)
    dataHeader =[[s01,s02,s03]]
    t1=Table(dataHeader,colWidths=(90*mm,29.5*mm,90*mm))
    t1.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),('TEXTCOLOR',(0,0),(-1,-1),black),('ALIGN',(0,0),(-1,-1),'RIGHT'),('VALIGN',(0,0),(-1,-1),'TOP'), ('LINEBELOW', (0,-1), (-1,-1), 1, colors.black),]))
    elements.append(t1)


    # if isStoreGlobal:

    try:
        detailsObj = contract.user.profile.details
        details = ast.literal_eval(detailsObj)
        gstVal = details['GST']
    except:
        gstVal = ''

    currencySymbol = appSettingsField.objects.filter(name='currencySymbol')
    if len(currencySymbol)>0:
        if currencySymbol[0].value == 'fa-usd':
            currency = 'USD'
        else:
            currency = 'INR'
    else:
        currency = 'USD'

    dataDetails = []


    # s11 =Paragraph("<para fontSize=8> <br/> "+ str(contract.user.first_name) +" "+ str(contract.user.last_name) +"<br/>102 Eden Park, 20 Vittal Mallya Rd.<br/>"+ str(contract.city) +" - "+ str(contract.pincode) +"<br/>"+ str(contract.state) +" <br/> "+ str(contract.country) +"<br/>"+ str(contract.mobileNo) +"<br/><br/><br/></para>",styles['Normal'])
    #
    #
    # s12 =Paragraph("<para> </para>",styles['Normal'])
    # s13 = Paragraph("<para alignment='right' fontSize=18> INVOICE </para>",styles['Normal'])
    s11 =Paragraph("<para fontSize=8>  <br/> <br/><b> Shipping Address </b> <br/> "+ str(contract.user.first_name) +" "+ str(contract.user.last_name) +"<br/>102 Eden Park, 20 Vittal Mallya Rd.<br/>"+ str(contract.city) +" - "+ str(contract.pincode) +"<br/>"+ str(contract.state) +" <br/> "+ str(contract.country) +"<br/>"+ str(contract.mobileNo) +"<br/><br/><br/></para>",styles['Normal'])
    s12 =Paragraph("<para fontSize=8>  <br/> <br/><b> Billing Address </b> <br/> "+ str(contract.user.first_name) +" "+ str(contract.user.last_name) +"<br/>102 Eden Park, 20 Vittal Mallya Rd.<br/>"+ str(contract.billingCity) +" - "+ str(contract.billingPincode) +"<br/>"+ str(contract.billingState) +" <br/> "+ str(contract.billingState) +"<br/>"+ str(contract.mobileNo) +"<br/><br/><br/></para>",styles['Normal'])
    s13 = Paragraph("<para alignment='right' fontSize=18> INVOICE </para>",styles['Normal'])
    dataDetails +=[[s11,s12,s13]]

    invoiceNumber = ''
    orderNumber = str(contract.pk)

    if isStoreGlobal:
        s21 =Paragraph("<para fontSize=8>Order Number : "+ orderNumber +"</para>",styles['Normal'])
    else:
        s21 =Paragraph("<para fontSize=8> GSTIN : "+ gstVal +"<br/>Order Number : "+ orderNumber +"</para>",styles['Normal'])

    s22 =Paragraph("<para> </para>",styles['Normal'])

    print datetime.datetime.now()
    invoiceDate = str(datetime.datetime.now().date())
    orderDate = str(contract.created.date())


    s23 = Paragraph("<para alignment='right' fontSize=8> Invoice Date: "+ invoiceDate +" <br/>Order Date: "+ orderDate +" </para>",styles['Normal'])
    dataDetails +=[[s21,s22,s23]]
    t2=Table(dataDetails,colWidths=(85.5*mm,85.5*mm,40*mm))
    t2.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),('TEXTCOLOR',(0,0),(-1,-1),black),('ALIGN',(0,0),(-1,-1),'RIGHT'),('VALIGN',(0,0),(-1,-1),'TOP'), ]))
    elements.append(t2)

    p101_01 =Paragraph("<para fontSize=8>Sl. no</para>",styles['Normal'])
    p101_02 =Paragraph("<para fontSize=8>Product</para>",styles['Normal'])
    p101_09 =Paragraph("<para fontSize=8>Qty</para>",styles['Normal'])
    p101_03 =Paragraph("<para fontSize=8>HSN</para>",styles['Normal'])
    p101_04 =Paragraph("<para fontSize=8>Price <span fontSize=6> ("+currency+")</span></para>",styles['Normal'])
    p101_05 =Paragraph("<para fontSize=8>SGST  <span fontSize=6> ("+currency+")</span></para>",styles['Normal'])
    p101_06 =Paragraph("<para fontSize=8>CGST  <span fontSize=6> ("+currency+")</span></para>",styles['Normal'])
    p101_07 =Paragraph("<para fontSize=8>Total GST  <span fontSize=6> ("+currency+")</span></para>",styles['Normal'])
    p101_08 =Paragraph("<para fontSize=8>Total Price  <span fontSize=6> ("+currency+")</span></para>",styles['Normal'])

    if isStoreGlobal:
        data1=[[p101_01,p101_02,p101_09,p101_04,p101_08]]
    else:
        data1=[[p101_01,p101_02,p101_09,p101_03,p101_04,p101_05,p101_06,p101_07,p101_08]]
    # p101_06 =Paragraph("<para fontSize=8>{0}</para>".format(amountTitle),styles['Normal']) IF ANY NON STATIC DATA GIIVE LIKE THIS

    rheights=1*[0.2*inch]
    if isStoreGlobal:
        t3=Table(data1,rowHeights=rheights,colWidths=(15*mm, 100*mm,25*mm , 35*mm, 35*mm))
    else:
        t3=Table(data1,rowHeights=rheights,colWidths=(15*mm, 50*mm,10*mm ,20*mm, 25*mm, 20*mm, 20*mm, 25*mm, 25*mm))
    t3.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),('TEXTCOLOR',(0,0),(-1,-1),black),('ALIGN',(0,0),(-1,-1),'LEFT'),('VALIGN',(0,0),(-1,-1),'TOP'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
    elements.append(t3)
    Serialid=0
    data2 = []
    totaldiscount = 0
    total = 0
    totalprice = 0
    promoAmount = 0
    discount = 0
    promoCode = ''
    subtotal = 0
    promoObj = Promocode.objects.all()
    if contract.promoCode:
        for p in promoObj:
            if str(p.name)==str(contract.promoCode):
                promoAmount = p.discount
                print  p.name,'aaaaaaaaaaaaaaaa'
                promoCode = p.name
    else:
        promoCode="None"

    for i in contract.orderQtyMap.all():
        Serialid+=1
        if i.prodSku == i.product.product.serialNo:
            print i.product.product.name, i.product.product.discount, i.priceDuringOrder
            price = i.priceDuringOrder - (i.product.product.discount * i.priceDuringOrder)/100
            price=round(price, 2)
            totalprice = i.qty*price
            totalprice=round(totalprice, 2)
            total=round(total, 2)
            if i.desc:
                desc =i.desc
            else:
                desc=""
            qtyData = i.product.product.howMuch
            if str(i.product.product.unit)=='Gram' or str(i.product.product.unit)=='gm':
                if qtyData >1000:
                    qtyValue = str(qtyData/1000) + ' Kg'
                else:
                    qtyValue = str(qtyData) + ' gm'
            elif str(i.product.product.unit)=='Millilitre' or str(i.product.product.unit)=='ml':
                if qtyData>1000:
                    qtyValue = str(qtyData/1000) + ' lt'
                else:
                    qtyValue = str(qtyData) + ' ml'
            elif str(i.product.product.unit)=='Size and Color' or str(i.product.product.unit)=='Size':
                if int(qtyData)==1:
                    qtyValue = 'XS'
                elif int(qtyData)==2:
                    qtyValue = 'S'
                elif int(qtyData)==2:
                    qtyValue = 'M'
                elif int(qtyData)==4:
                    qtyValue = 'L'
                elif int(qtyData)==5:
                    qtyValue = 'XL'
                elif int(qtyData)==6:
                    qtyValue = 'XL'
                else:
                    qtyValue = qtyData
            else:
              qtyValue = qtyData
            prodName = str(i.product.product.name) + ' '  + str(qtyValue) + ' ' +str(desc)
            if i.product.product.productMeta is None:
                gstAmount = ''
                hsnCode = ''
                gstAmount2 = ''
            else:
                totalprice = totalprice + i.gstAmount
                gstAmount = i.gstAmount
                gstAmount2 = gstAmount/2
                hsnCode = str(i.product.product.productMeta.code)
            total+=totalprice
        else:
            prodData = ProductVerient.objects.get(sku = i.prodSku)
            price = prodData.discountedPrice
            price=round(price, 2)
            totalprice = i.qty*price
            totalprice=round(totalprice, 2)
            total=round(total, 2)
            if i.desc:
                desc =i.desc
            else:
                desc=""
            qtyData = i.product.product.howMuch * prodData.unitPerpack
            if str(i.product.product.unit)=='Gram' or str(i.product.product.unit)=='gm':
                if qtyData >1000:
                    qtyValue = str(qtyData/1000) + ' Kg'
                else:
                    qtyValue = str(qtyData) + ' gm'
            elif str(i.product.product.unit)=='Millilitre' or str(i.product.product.unit)=='ml':
                if qtyData>1000:
                    qtyValue = str(qtyData/1000) + ' lt'
                else:
                    qtyValue = str(qtyData) + ' ml'
            elif str(i.product.product.unit)=='Size and Color' or str(i.product.product.unit)=='Size':
                if int(qtyData)==1:
                    qtyValue = 'XS'
                elif int(qtyData)==2:
                    qtyValue = 'S'
                elif int(qtyData)==2:
                    qtyValue = 'M'
                elif int(qtyData)==4:
                    qtyValue = 'L'
                elif int(qtyData)==5:
                    qtyValue = 'XL'
                elif int(qtyData)==6:
                    qtyValue = 'XL'
                else:
                    qtyValue = qtyData
            else:
              qtyValue = qtyData


            prodName = str(i.product.product.name) + ' ' + str(qtyValue)+ ' ' +str(desc)

            if i.product.product.productMeta is None:
                gstAmount = ''
                hsnCode = ''
                gstAmount2 = ''
            else:
                totalprice = totalprice + i.gstAmount
                gstAmount = i.gstAmount
                gstAmount2 = gstAmount/2
                hsnCode = str(i.product.product.productMeta.code)
            total+=totalprice

        qty = i.qty
        subtotal += i.totalAmount
        p111_01 =Paragraph("<para fontSize=8>"+ str(Serialid) +"</para>",styles['Normal'])
        p111_02 =Paragraph("<para fontSize=8> "+ str(prodName) +" </para>",styles['Normal'])
        p111_09 =Paragraph("<para fontSize=8> "+ str(qty) +" </para>",styles['Normal'])
        p111_03 =Paragraph("<para fontSize=8>" + str(hsnCode) +"</para>",styles['Normal'])
        p111_04 =Paragraph("<para fontSize=8> "+ str(price) +"</para>",styles['Normal'])
        p111_05 =Paragraph("<para fontSize=8>"+ str(gstAmount2)+"</para>",styles['Normal'])
        p111_06 =Paragraph("<para fontSize=8>"+ str(gstAmount2)+"</para>",styles['Normal'])
        p111_07 =Paragraph("<para fontSize=8>"+ str(gstAmount)+" </para>",styles['Normal'])
        p111_08 =Paragraph("<para fontSize=8  alignment='right'>"+ str(i.totalAmount) +"</para>",styles['Normal'])

        if isStoreGlobal:
            data2.append([p111_01,p111_02,p111_09,p111_04,p111_08])
        else:
            data2.append([p111_01,p111_02,p111_09,p111_03,p111_04,p111_05,p111_06,p111_07,p111_08])


    shippingCharges = round(contract.shippingCharges,2)
    subtotal = round(subtotal,2)
    grandTotal=round(contract.totalAmount, 2)
    if isStoreGlobal:
        t4=Table(data2,colWidths=(15*mm, 100*mm,25*mm , 35*mm, 35*mm))
    else:
        t4=Table(data2,colWidths=(15*mm, 50*mm,10*mm ,20*mm, 25*mm, 20*mm, 20*mm, 25*mm, 25*mm))

    t4.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),('TEXTCOLOR',(0,0),(-1,-1),black),('ALIGN',(0,0),(-1,-1),'LEFT'),('VALIGN',(0,0),(-1,-1),'TOP')]))
    elements.append(t4)
    p121_01 =Paragraph("<para fontSize=8></para>",styles['Normal'])
    p121_02 =Paragraph("<para fontSize=8 alignment='left'> Subtotal</para>",styles['Normal'])
    p121_03 =Paragraph("<para fontSize=8  alignment='right'> <span fontSize=7> "+currency+ "</span>  "+str(subtotal) +"</para>",styles['Normal'])
    data3=[[p121_01,p121_02,p121_03]]

    p131_01 =Paragraph("<para fontSize=8></para>",styles['Normal'])
    p131_02 =Paragraph("<para fontSize=8 alignment='left'>Shipping Charges</para>",styles['Normal'])
    p131_03 =Paragraph("<para fontSize=8  alignment='right'> <span fontSize=7> "+ currency+ "</span> "+ str(shippingCharges) +"</para>",styles['Normal'])
    data3+=[[p131_01,p131_02,p131_03]]

    # p132_01 =Paragraph("<para fontSize=8></para>",styles['Normal'])
    # p132_02 =Paragraph("<para fontSize=8 alignment='left'>Promo Code Discount</para>",styles['Normal'])
    # p132_03 =Paragraph("<para fontSize=8  alignment='right'>"+ str(promoAmount) +"</para>",styles['Normal'])
    # data3+=[[p132_01,p132_02,p132_03]]

    t5=Table(data3,colWidths=(140*mm, 30*mm,40*mm))
    t5.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),('TEXTCOLOR',(0,0),(-1,-1),black),('ALIGN',(0,0),(-1,-1),'LEFT'),('VALIGN',(0,0),(-1,-1),'TOP')]))
    elements.append(t5)

    p141_01 =Paragraph("<para fontSize=8></para>",styles['Normal'])
    p141_02 =Paragraph("<para fontSize=8 alignment='left'>Total </para>",styles['Normal'])
    p141_03 =Paragraph("<para fontSize=8  alignment='right'> <span fontSize=7> "+ currency + "</span> "+ str(grandTotal) +"</para>",styles['Normal'])
    data4=[[p141_01,p141_02,p141_03]]
    t6=Table(data4,colWidths=(140*mm, 30*mm,40*mm))
    t6.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),('TEXTCOLOR',(0,0),(-1,-1),black),('ALIGN',(0,0),(-1,-1),'LEFT'),('VALIGN',(0,0),(-1,-1),'TOP'), ('LINEBELOW', (1,0), (-1,-1), 0.5, colors.black),('LINEABOVE', (1,0), (-1,-1), 0.5, colors.black),]))
    elements.append(t6)





    doc.build(elements, canvasmaker=PageCanvas)

class DownloadInvoiceAPI(APIView):
    renderer_classes = (JSONRenderer,)
    def get(self, request, format=None):
        response = HttpResponse(content_type='application/pdf')
        o = Order.objects.get(pk=request.GET['value'])
        # invoice = 'bni'
        # if invoice=='bni':
        #     template_path = 'invoice.html'
        #     context = {'myvar': {'name':'vikas'}}
        #     response = HttpResponse(content_type='application/pdf')
        #     response['Content-Disposition'] = 'attachment; filename="report.pdf"'
        #     template = get_template(template_path)
        #     html = template.render(context)
        #     pisaStatus = pisa.CreatePDF(
        #        html, dest=response, link_callback=link_callback)
        #     if pisaStatus.err:
        #        return HttpResponse('We had some errors <pre>' + html + '</pre>')
        #     return response
        # else:
        print o
        response['Content-Disposition'] = 'attachment; filename="invoice%s_%s.pdf"' % (
             datetime.datetime.now(pytz.timezone('Asia/Kolkata')).year, o.pk)
        if re.search('BNI', globalSettings.SEO_TITLE, re.IGNORECASE):
            print 'call bni'
            generateInvoiceBni(response, o, request)
        else:
            genInvoice(response, o, request)
        f = open(os.path.join(globalSettings.BASE_DIR, 'media_root/invoice%s_%s.pdf' %
                              ( datetime.datetime.now(pytz.timezone('Asia/Kolkata')).year, o.pk)), 'wb')
        f.write(response.content)
        f.close()
        return response


class RatingViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny , )
    queryset = Rating.objects.all()
    serializer_class = RatingSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['productDetail',]

class SupportFeedViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny , )
    queryset = SupportFeed.objects.all()
    serializer_class = SupportFeedSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['status',]

# class suggestedItemAPI(APIView):
#     renderer_classes = (JSONRenderer,)
#     def get(self, request, format=None):
#         count=0
#         listObj = listing.objects.all()
#         for i in listObj:
#             average=0
#             ratingObj = Rating.objects.filter(productDetail=i.pk)
#             if ratingObj.count()!=0:
#                 rating = 0.0
#                 for j in ratingObj:
#                     rating+= j.rating
#                 average=round(rating/ratingObj.count(),2)
#                 print average,'avvvvvvvvvvvv'
#         # return(status.HTTP_200_OK)

from datetime import timedelta
from django.db.models import Sum
class OnlineSalesGraphAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.IsAuthenticated ,)
    def post(self , request , format = None):
        totalCollections=0
        if "date" in request.data:
            # one day sale
            d = datetime.datetime.strptime(request.data["date"], '%Y-%m-%dT%H:%M:%S.%fZ')
            print d,'dateeeeeeeeeeeeeeee'
            order = Order.objects.filter(created__range = (datetime.datetime.combine(d, datetime.time.min), datetime.datetime.combine(d, datetime.time.max)))
            custs = User.objects.filter(date_joined__range= (datetime.datetime.combine(d, datetime.time.min), datetime.datetime.combine(d, datetime.time.max)))
            orderQty = OrderQtyMap.objects.filter(updated__range = (datetime.datetime.combine(d, datetime.time.min), datetime.datetime.combine(d, datetime.time.max)))
        else:
            frm = datetime.datetime.strptime(request.data["from"], '%Y-%m-%dT%H:%M:%S.%fZ')
            to = datetime.datetime.strptime(request.data["to"], '%Y-%m-%dT%H:%M:%S.%fZ')
            order = Order.objects.filter(created__range=(datetime.datetime.combine(frm, datetime.time.min), datetime.datetime.combine(to, datetime.time.max)))
            orderQty = OrderQtyMap.objects.filter(updated__range = (datetime.datetime.combine(frm, datetime.time.min), datetime.datetime.combine(to, datetime.time.max)))
            custs = User.objects.filter(date_joined__range = (datetime.datetime.combine(frm, datetime.time.min), datetime.datetime.combine(to, datetime.time.max)))

        totalSales = order.aggregate(Sum('totalAmount')) if order.count() > 0 else {'totalAmount__sum':0}
        if 'totalAmount__sum' in totalSales and type(totalSales['totalAmount__sum']) == float:
            totalSales['totalAmount__sum'] = round(totalSales['totalAmount__sum'],2)
        for i in orderQty:
            if str(i.status) == 'delivered':
                price = i.product.product.price - (i.product.product.discount * i.product.product.price)/100
                orderD = Order.objects.filter(orderQtyMap=i.pk)
                for j in orderD:
                        if j.promoCode!=None:
                            promo = Promocode.objects.filter(name__iexact=j.promoCode)
                            for p in promo:
                                promocode = p.discount
                                priceVal = price-(promocode * price)/100
                                totalCollections += priceVal
                            else:
                                totalCollections += price
            elif str(i.status) != 'delivered':
                orderD = Order.objects.filter(orderQtyMap=i.pk)
                for j in orderD:
                    if str(j.paymentMode) == 'card':
                        price = i.product.product.price - (i.product.product.discount * i.product.product.price)/100
                        if j.promoCode!=None:
                            promo = Promocode.objects.filter(name__iexact=j.promoCode)
                            for p in promo:
                                promocode = p.discount
                                priceVal = price-(promocode * price)/100
                                totalCollections += priceVal
                        else:
                            totalCollections += price
        totalCollections = round(totalCollections, 2)
        sales =  order.count()
        custCount = custs.count()


        last_month = datetime.datetime.now() - timedelta(days=30)

        data = (Order.objects.all()
            .extra(select={'created': 'date(created)'})
            .values('created')
            .annotate(sum=Sum('totalAmount')))


        # return Response({"totalSales" : totalSales , "totalCollections" : totalCollections ,  "sales" : sales , "custCount" : custCount , "trend" : data},status=status.HTTP_200_OK)
        return Response({"totalSales" : totalSales , "totalCollections" : totalCollections ,  "sales" : sales , "custCount" : custCount , "trend" : data},status=status.HTTP_200_OK)

class GenericPincodeViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny ,)
    # queryset = GenericPincode.objects.all()
    serializer_class = genericPincodeSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['pincode','state','city']
    def get_queryset(self):
        toReturn = GenericPincode.objects.all()
        if 'pincode' in self.request.GET:
            toReturn = toReturn.filter(pincode__iexact=self.request.GET['pincode'])
        return toReturn

class GenericImageViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny ,)
    queryset = GenericImage.objects.all()
    serializer_class = genericImageSerializer
    # filter_backends = [DjangoFilterBackend]
    # filter_fields = ['pincode','state','city']
import traceback
class UpdateCartAPIView(APIView):
    permission_classes = (permissions.IsAuthenticated ,)
    def post(self, request, format=None):
        try:
            c = Cart.objects.get(product__id = request.data['product'] , user = request.user)
            c.qty  = request.data['qty']
            c.save()
            return Response(status=status.HTTP_200_OK)
        except:
            traceback.print_exc()
            return Response(status = status.HTTP_404_NOT_FOUND )



class BulklistingCreationAPIView(APIView):
    permission_classes = (permissions.IsAuthenticated , isAdmin)
    def post(self, request, format=None):
        wb = load_workbook(filename = BytesIO(request.FILES['xl'].read()))
        ws = wb.worksheets[0]
        row_count = ws.max_row+1
        column_count = ws.max_column
        for i in range(2, row_count):
            images =[]
            genericObj=''
            name = ws['A' + str(i)].value
            serialNo = ws['B' + str(i)].value
            price = ws['C' + str(i)].value
            serialId = ws['D' + str(i)].value
            unit = ws['E' + str(i)].value
            howMuch = ws['F' + str(i)].value
            grossWeight = ws['K' + str(i)].value
            send = Product(name=name, serialNo=serialNo, price=price,serialId=serialId,unit=unit,howMuch=howMuch,user=request.user,grossWeight=grossWeight)
            send.save()
            cat = ws['J' + str(i)].value
            try:
                genericObj = genericProduct.objects.get(name__iexact = cat)
                parentType = genericObj
            except:
                genericName = cat
                visual = 'ecommerce/pictureUploads/' + str(cat) + '_small.jpg'
                bannerImage = 'ecommerce/GenericProductBanner/' + str(cat) + '_big.jpg'
                genericSend = genericProduct(name=genericName,visual=visual,bannerImage=bannerImage)
                genericSend.save()
                parentType = genericSend
            imageOne = ws['G' + str(i)].value
            attachmentOne = 'ecommerce/pictureUploads/' +str(imageOne)
            imageOneSend = media(user=request.user,attachment = attachmentOne,mediaType='image' )
            imageOneSend.save()
            images.append(imageOneSend)
            imageTwo = ws['H' + str(i)].value
            attachmentTwo = 'ecommerce/pictureUploads/' +str(imageTwo)
            imageTwoSend = media(user=request.user,attachment = attachmentTwo,mediaType='image' )
            imageTwoSend.save()
            images.append(imageTwoSend)
            imageThree = ws['I' + str(i)].value
            attachmentThree = 'ecommerce/pictureUploads/' +str(imageThree)
            imageThreeSend =media(user=request.user,attachment = attachmentThree,mediaType='image' )
            imageThreeSend.save()
            images.append(imageThreeSend)
            try:
                productIndex = ws['L' + str(i)].value
                listingSend = listing(parentType=parentType, product=send,user=request.user,productIndex=productIndex)
            except:
                listingSend = listing(parentType=parentType, product=send,user=request.user)
            listingSend.save()
            for i in images:
                listingSend.files.add(i)
            listingSend.save()
        return Response(status = status.HTTP_200_OK)

from paypal.standard.forms import PayPalPaymentsForm
def paypal_return_view(request):
    orderObj = Order.objects.filter(user = request.user).last()
    return updateAndProcessOrder(orderObj.id , orderObj.totalAmount)
    # orderObj.paidAmount = orderObj.totalAmount
    # orderObj.approved = True
    # orderObj.save()
    #
    # # send the email here
    # value = []
    # totalPrice = 0
    # promoAmount = 0
    # total=0
    # price = 0
    # grandTotal = 0
    # promoObj = Promocode.objects.all()
    # for p in promoObj:
    #     if str(p.name)==str(orderObj.promoCode):
    #         promoAmount = p.discount
    # print promoAmount
    # a = '#'
    # docID = str(a) + str(orderObj.pk)
    # for i in orderObj.orderQtyMap.all():
    #     if i.prodSku == i.product.product.serialNo:
    #         price = i.product.product.price - (i.product.product.discount * i.product.product.price)/100
    #         price=round(price, 2)
    #         totalPrice=i.qty*price
    #         totalPrice=round(totalPrice, 2)
    #         total+=totalPrice
    #         total=round(total, 2)
    #         value.append({ "productName" : i.product.product.name,"qty" : i.qty , "amount" : totalPrice,"price":price})
    #     else:
    #         prodData = ProductVerient.objects.get(sku = i.prodSku)
    #         price = prodData.discountedPrice
    #         price=round(price, 2)
    #         totalPrice=i.qty*price
    #         totalPrice=round(totalPrice, 2)
    #         total+=totalPrice
    #         total=round(total, 2)
    #         value.append({ "productName" : i.product.product.name,"qty" : i.qty , "amount" : totalPrice,"price":price})
    # grandTotal=total-(promoAmount * total)/100
    # grandTotal=round(grandTotal, 2)
    # request.user.cartItems.all().delete()
    # try:
    #     isStoreGlobal = appSettingsField.objects.filter(name='isStoreGlobal')[0].flag
    # except:
    #     isStoreGlobal = False
    # try:
    #     companyAddress = appSettingsField.objects.filter(name='companyAddress')[0].value
    #     cleanr = re.compile('<.*?>')
    #     companyAddress = re.sub(cleanr, '', companyAddress)
    # except:
    #     companyAddress = ''
    #
    # try:
    #     gstValue = appSettingsField.objects.filter(name='cstNo')[0].value
    # except:
    #     gstValue = ''
    #
    # if orderObj.user.email:
    #     ctx = {
    #         'heading' : "Invoice Details",
    #         'recieverName' : orderObj.user.first_name  + " " +orderObj.user.last_name ,
    #         'linkUrl': globalSettings.BRAND_NAME,
    #         'sendersAddress' : globalSettings.SEO_TITLE,
    #         # 'sendersPhone' : '122004',
    #         'grandTotal':grandTotal,
    #         'total': total,
    #         'value':value,
    #         'docID':docID,
    #         'data':orderObj,
    #         'promoAmount':promoAmount,
    #         'linkedinUrl' : lkLink,
    #         'fbUrl' : fbLink,
    #         'twitterUrl' : twtLink,
    #         'isStoreGlobal':isStoreGlobal,
    #         'companyAddress':companyAddress,
    #         'gstValue':gstValue
    #     }
    #     print ctx
    #     contactData = []
    #     email_body = get_template('app.ecommerce.emailDetail.html').render(ctx)
    #     email_subject = "Order Details"
    #     email_to = []
    #     email_to.append(str(orderObj.user.email))
    #     email_cc = []
    #     email_bcc = []
    #     send_email(email_body,email_to,email_subject,email_cc,email_bcc,'html')
    # return redirect("/checkout/cart?action=success&orderid=" + str(orderObj.pk))


def paypal_cancel_view(request):
    return redirect("/checkout/cart?action=retry")

def paypalPaymentInitiate(request):
    # What you want the button to do.
    orderid = request.GET['orderid']
    orderObj = Order.objects.get(pk=orderid)
    paypal_dict = {
        "business": globalSettings.PAYPAL_RECEIVER_EMAIL,
        "amount": str(orderObj.totalAmount),
        "item_name": "BNI marchandise",
        "invoice": orderObj.pk,
        'currency_code': 'USD',
        "notify_url": request.build_absolute_uri(reverse('paypal-ipn')),
        "return": request.build_absolute_uri(reverse('paypal_return_view')),
        "cancel_return": request.build_absolute_uri(reverse('paypal_cancel_view')),
        "custom": "premium_plan",  # Custom command to correlate to some function later (optional)
    }

    # Create the instance.
    form = PayPalPaymentsForm(initial=paypal_dict)
    context = {"form": form}
    return render(request, "paypal.payment.html", context)

import uuid
def payuPaymentInitiate(request):
    # What you want the button to do.
    orderid = request.GET['orderid']
    orderObj = Order.objects.get(pk=orderid)



    hashSequence = "key|txnid|amount|productinfo|firstname|email|udf1|udf2|udf3|udf4|udf5|udf6|udf7|udf8|udf9|udf10";

    hash_string = '';
    hashVarsSeq = hashSequence.split('|');
    trxnID = str(uuid.uuid4()).split('-')[0]
    posted = {"key"  : globalSettings.PAYU_MERCHANT_KEY ,
        "txnid" : trxnID ,
        "amount" : str(orderObj.totalAmount),
        "productinfo" : "Sterling select products",
        "firstname" : orderObj.user.first_name,
        "email" : orderObj.user.email}

    for hvs in hashVarsSeq:
        try:
            hash_string += posted[hvs];
        except:
            hash_string += ''

        hash_string += '|'

    orderObj.paymentRefId = trxnID
    orderObj.save()

    hash_string += globalSettings.PAYU_MERCHANT_SALT
    print hash_string
    hashh = hashlib.sha512(hash_string).hexdigest()
    print "hashh : " , hashh
    formData =  {
        "action" : payu_url(),
        "key": globalSettings.PAYU_MERCHANT_KEY,
        "txnid": trxnID,
        "amount" : str(orderObj.totalAmount),
        "productinfo" : "Sterling select products",
        "firstname" : orderObj.user.first_name,
        "email" : orderObj.user.email,
        "phone" : '9702438730',
        "surl" :  globalSettings.SITE_ADDRESS +'/payUPaymentResponse/',
        "furl" : globalSettings.SITE_ADDRESS +'/payUPaymentResponse/',
        "hash" : hashh
    }

    print formData

    return render(request , 'payu.payment.html' , formData)

from django.views.decorators.csrf import csrf_exempt, csrf_protect


def updateAndProcessOrder(orderID , amnt, referenceId=None):
    print 'in updateAndProcessOrderupdateAndProcessOrder'
    orderObj = Order.objects.get(id = orderID)
    orderObj.paidAmount = amnt
    orderObj.approved = True
    orderObj.paymentStatus = True
    orderObj.save()
    value = []
    totalPrice = 0
    promoAmount = 0
    total=0
    price = 0
    grandTotal = 0
    promoObj = Promocode.objects.all()
    for p in promoObj:
        if str(p.name)==str(orderObj.promoCode):
            promoAmount = p.discount
    print promoAmount
    a = '#'
    docID = str(a) + str(orderObj.pk)
    try:
        isStoreGlobal = appSettingsField.objects.filter(name='isStoreGlobal')[0].flag
    except:
        isStoreGlobal = False
    for i in orderObj.orderQtyMap.all():
        i.paidAmount = i.totalAmount
        i.save()
        if i.prodSku == i.product.product.serialNo:
            price = i.product.product.price - (i.product.product.discount * i.product.product.price)/100
            price=round(price, 2)
            totalPrice=i.qty*price
            totalPrice=round(totalPrice, 2)

            total=round(total, 2)
            if not isStoreGlobal:
                if i.product.product.productMeta is not None:
                    gst = (i.product.product.productMeta.taxRate * price) /100
                    totalPrice =totalPrice + gst
                else:
                    gst = 0
                    totalPrice = totalPrice + gst
            else:
                gst = 0
            total+=totalPrice
            value.append({ "productName" : i.product.product.name,"qty" : i.qty , "amount" : i.totalAmount,"price":price,"gst":i.gstAmount})
        else:
            prodData = ProductVerient.objects.get(sku = i.prodSku)
            price = prodData.discountedPrice
            price=round(price, 2)
            totalPrice=i.qty*price
            totalPrice=round(totalPrice, 2)
            total=round(total, 2)
            if not isStoreGlobal:
                if i.product.product.productMeta is not None:
                    gst = (i.product.product.productMeta.taxRate * price) /100
                    totalPrice =totalPrice + gst
                else:
                    gst = 0
                    totalPrice = totalPrice + gst
            else:
                gst = 0
            total+=totalPrice
            value.append({ "productName" : i.product.product.name,"qty" : i.qty , "amount" : i.totalAmount,"price":price,"gst":i.gstAmount})
    shippingCharges = round(orderObj.shippingCharges,2)
    grandTotal = round(orderObj.totalAmount, 2)
    orderObj.user.cartItems.all().delete()
    print orderObj.user.email, 'email'
    print 'semndddddddddddddddd emaillllllllll'

    try:
        currencyTyp = appSettingsField.objects.filter(name='currencySymbol')[0].value
        if currencyTyp == 'fa-usd':
            currencyVal = 'USD'
        else:
            currencyVal = 'INR'
    except:
        currencyVal = ''

    try:
        companyAddress = appSettingsField.objects.filter(name='companyAddress')[0].value
    except:
        companyAddress = ''
    try:
        gstValue = appSettingsField.objects.filter(name='cstNo')[0].value
    except:
        gstValue = ''

    if orderObj.user.email:
        ctx = {
            'heading' : "Invoice Details",
            'recieverName' : orderObj.user.first_name  + " " +orderObj.user.last_name ,
            'linkUrl': globalSettings.BRAND_NAME,
            'sendersAddress' : globalSettings.SEO_TITLE,
            # 'sendersPhone' : '122004',
            'grandTotal':grandTotal,
            'total': total,
            'value':value,
            'docID':docID,
            'data':orderObj,
            'promoAmount':promoAmount,
            'linkedinUrl' : lkLink,
            'fbUrl' : fbLink,
            'twitterUrl' : twtLink,
            'isStoreGlobal':isStoreGlobal,
            'companyAddress':companyAddress,
            'gstValue':gstValue,
            'currencyVal':currencyVal,
            'shippingCharges':shippingCharges
        }
        email_body = get_template('app.ecommerce.emailDetail.html').render(ctx)
        email_subject = 'Order Placed'
        email_to = []
        email_to.append(str(orderObj.user.email))
        email_cc = []
        email_bcc = []
        send_email(email_body,email_to,email_subject,email_cc,email_bcc,'html')
    return redirect("/checkout/cart?action=success&orderid=" + str(orderObj.pk))


@csrf_exempt
def payUPaymentResponse(request):
    if request.method == 'POST' and request.POST['status'] == 'success':
        return updateAndProcessOrder(request.POST['txnid'] ,  request.POST['amount'])
    else:
        return redirect("/checkout/cart?action=retry")

@csrf_exempt
def ebsPaymentResponse(request):
    # return updateAndProcessOrder(213 ,  800)
    if request.method == 'POST' and 'Successful' in request.POST['ResponseMessage']:
        #ResponseCode=0&ResponseMessage=Transaction+Successful&DateCreated=2018-12-30+12%3A03%3A28&PaymentID=118878521&MerchantRefNo=40&Amount=2.00&Mode=LIVE&BillingName=Admin&BillingAddress=ABC+%2C+kudlu&BillingCity=Bengaluru&BillingState=Karnataka&BillingPostalCode=560068&BillingCountry=IND&BillingPhone=9702438730&BillingEmail=pkyisky%40gmail.com&DeliveryName=&DeliveryAddress=&DeliveryCity=&DeliveryState=&DeliveryPostalCode=&DeliveryCountry=&DeliveryPhone=&Description=test+&IsFlagged=NO&TransactionID=355107915&PaymentMethod=1161&RequestID=84015197&SecureHash=1F417B260A360D23DE7B7F3C9E054296
        return updateAndProcessOrder(request.POST['MerchantRefNo'] , request.POST['Amount'])
    else:
        return redirect("/checkout/cart?action=retry")


class GetInStockAPIView(APIView):
    # permission_classes = (permissions.IsAuthenticated , isAdmin)
    def get(self, request, format=None):
        print request.GET['store'] ,  request.GET['product_var'],request.GET['product_id'],'eeeeeeeeeeeee'
        store = None
        product_var = None
        stock=0
        if request.GET['store']!='undefined':
            store = int(request.GET['store'])
        if  request.GET['product_var']!='undefined':
            product_var = int(request.GET['product_var'])
        storeData = StoreQty.objects.get(store=store,product = request.GET['product_id'],productVariant=product_var)
        if storeData:
            stock = storeData.quantity
        else:
            stock=0
        print stock , 'sttccccccccccccccccckkkkkkkkkkk'
        stockData = { 'stock':stock,'product':request.GET['product_id'],'product_var':request.GET['product_var'],'store':request.GET['store']}
        return Response(stockData,status = status.HTTP_200_OK)


class SearchCountryAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny , )
    def get(self , request , format = None):
        if 'getCountryCode' in request.GET:
            toReturn = []
            try:
                c = Countries.objects.filter(name__iexact = request.GET['getCountryCode'])
                if len(c)>0:
                    toReturn.append(c[0].sortname)
                else:
                    toReturn.append('US')
                return Response(toReturn, status = status.HTTP_200_OK)
            except:
                toReturn = ['US']
                return Response(toReturn, status = status.HTTP_200_OK)

        if 'country' in request.GET:
            print 'state' , request.GET['country'], request.GET['query']
            states = States.objects.filter(name__icontains=request.GET['query'], country_id = request.GET['country'])
            print len(states)
            return Response(list(states.values())[:10], status = status.HTTP_200_OK)
        elif 'state' in request.GET:
            city = Cities.objects.filter(name__icontains=request.GET['query'], state_id = request.GET['state'])
            return Response(list(city.values())[:10], status = status.HTTP_200_OK)
        else:
            countries = Countries.objects.filter(Q(name__icontains=request.GET['query']) | Q(sortname__icontains=request.GET['query']))
            return Response(list(countries.values())[:10], status = status.HTTP_200_OK)

from create_shipment import createShipment

class CreateShipmentAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny , )
    def get(self , request , format = None):
        print request.GET ,'ddddddddd'
        order = Order.objects.get(pk = int(request.GET['orderPk']))
        recipientName =  order.user.first_name + ' ' + order.user.last_name
        print type(order.mobileNo) ,'hhhhhhhhhhhh'
        try:
            recipientPhone = int(order.mobileNo)
        except :
            recipientPhone = 9999
        recipientAddress = str(order.landMark) +" "+ str(order.street)
        city = order.city
        state = order.stateCode
        country = order.countryCode
        pincode = str(order.pincode)
        weight = float(request.GET['totalWeight']) * 2.20462
        awbPath , trackingID = createShipment(recipientName , recipientName , recipientPhone , [recipientAddress] ,  city, state , pincode , country,  weight)
        return Response({'awbPath':awbPath,'trackingID':trackingID,'courierName':'Fedex'}, status = status.HTTP_200_OK)

# from django.core.files import File
#
# class AddImageAPI(APIView):
#     renderer_classes = (JSONRenderer,)
#     permission_classes = (permissions.AllowAny , )
#     def get(self , request , format = None):
#         print 'GGGGGGGGGGGGGGGGGGGGGGGG'
#         allObj = Countries.objects.all()
#         for i in allObj:
#             name = i.sortname.lower()
#             try:
#                 img = open(os.path.join(globalSettings.BASE_DIR, 'media_root/Flags/'+name +'.png'))
#                 print File(img)
#                 i.flag.save(name+'.png', File(img))
#             except Exception as e:
#                 img = open(os.path.join(globalSettings.BASE_DIR, 'media_root/Flags/noImage.png'))
#                 i.flag.save(name+'.png', File(img))
#         return Response({'awbPath':12}, status = status.HTTP_200_OK)

class CountryViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny , )
    queryset = Countries.objects.all()
    serializer_class = CountrySerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['name']

class UserProfileSettingAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.IsAuthenticated ,)
    def get(self , request , format = None):
        print request.GET ,'gggggggggggggggggggggggggg'
        if 'user' in request.GET:
            user = User.objects.get(pk=request.GET['user'])
            prof = profile.objects.get(pk = user.profile.pk)
            firstName = user.first_name
            lastName = user.last_name
            email = user.email
            # pobj = profile.objects.get(pk=user.profile.pk)
            if prof.details is not None:
                if type(prof.details)==unicode:
                    details = ast.literal_eval(prof.details)
                else:
                    details = prof.details
            else:
                details = ''
            print details
            if 'GST' in details:
                print 'true'
                gst= details['GST']
                isGST = True
            else:
                print 'false'
                gst = ''
                isGST = False
            try:
                mobile = prof.mobile
            except:
                mobile =''
            return Response({'firstName':firstName,'lastName':lastName,'email':email,'mobile':mobile,'gst':gst,'isGST':isGST}, status = status.HTTP_200_OK)
    def post(self , request , format = None):
        if 'user' in request.data:
            user = User.objects.get(pk=request.data['user'])

            user.first_name = request.data['firstName']
            user.last_name = request.data['lastName']
            user.email = request.data['email']
            user.save()
            pobj = profile.objects.get(pk=user.profile.pk)
            # pobj.email = email
            pobj.mobile = request.data['mobile']
            toReturn = {'firstName':user.first_name,'lastName':user.last_name,'email':user.email,'mobile':pobj.mobile}
            if 'gst' in request.data:
                toReturn['gst'] = request.data['gst']
                if pobj.details is not None:
                    details = ast.literal_eval(pobj.details)
                    if 'GST' in details:
                        details['GST'] = request.data['gst']
                    else:
                        details['GST'] = request.data['gst']
                    pobj.details = details
                    pobj.save()
                else:
                    details = {}
                    details['GST'] = request.data['gst']
                    pobj.details = str(details)
                    pobj.save()
            pobj.save()
            toReturn['passwordChanged'] = False
            if 'oldPassword' in request.data:
                if authenticate(username = user.username , password = request.data['oldPassword']) is not None:
                    user.set_password(request.data['newPassword'])
                    user.save()
                    toReturn['passwordChanged'] = True
                else:
                    raise PermissionDenied(detail={'PARAMS':'Password missmatch'})
            return Response(toReturn, status = status.HTTP_200_OK)

class ReportsDataAPI(APIView):
    def get(self, request, format=None):
        print request.GET
        toRet = {}
        if 'typ' in request.GET:
            if request.GET['typ'] == 'sales':
                salesDownloadData = [['Name','Amount Received','GST Collected']]
                toReturn = []
                fdt = datetime.datetime.strptime(str(request.GET['fdt']),'%Y-%m-%d').date()
                sdt = datetime.datetime.strptime(str(request.GET['sdt']),'%Y-%m-%d').date()
                sdt = sdt+timedelta(days=1)
                print fdt,sdt,'dttttttttttttttt'
                qtyObjs = Order.objects.filter(orderQtyMap__updated__range=(fdt,sdt)).distinct()
                print qtyObjs
                for i in qtyObjs:
                    data = {}
                    tGst = 0
                    tAR = 0
                    data['name'] = 'ECOM_' + str(i.pk)
                    objs = i.orderQtyMap.filter(updated__range=(fdt,sdt))
                    for j in objs:
                        tAR += j.paidAmount
                        tGst += j.gstAmount
                    data['receivedAmount'] = tAR
                    if tAR==0:
                        tGst = 0
                    data['gstCollected'] = tGst
                    toReturn.append(data)
                    salesDownloadData.append(['ECOM_'+str(i.pk),tAR,tGst])

                invObjs = Invoice.objects.filter(updated__range=(fdt,sdt))
                print invObjs
                for i in invObjs:
                    data = {}
                    data['name'] = 'POS_' + str(i.pk)
                    data['receivedAmount'] = i.amountRecieved
                    tGst = 0
                    productsData = json.loads(i.products)
                    for j in productsData:
                        pp = j['data']['product']['price']
                        print j['data']['product']['productMeta']
                        if j['data']['product']['productMeta']:
                            print j['data']['product']['productMeta']
                            tx = j['data']['product']['productMeta']['taxRate']
                            gst = round(((pp*tx)/100)*i['quantity'])
                            tGst += gst
                    data['gstCollected'] = tGst
                    toReturn.append(data)
                    salesDownloadData.append(['POS_'+str(i.pk),i.amountRecieved,tGst])

                if 'download' in request.GET:
                    return ExcelResponse(salesDownloadData, 'SalesReport','Sales Details')

                return Response(toReturn,status = status.HTTP_200_OK)
            else:
                deliveryDownloadData = [['Name','Completed','Ongoing','COD','CARD','Total']]
                dt = datetime.datetime.strptime(str(request.GET['fdt']),'%Y-%m-%d').date()
                print dt,'dttttttttttttttt'
                qtyObjs = OrderQtyMap.objects.filter(updated__icontains=dt)
                invObjs = Invoice.objects.filter(updated__icontains=dt)
                qtyObjsUsersPks = list(qtyObjs.filter(orderBy__isnull=False).values_list('orderBy',flat=True).distinct())
                invObjsUsersPks = list(invObjs.filter(orderBy__isnull=False).values_list('orderBy',flat=True).distinct())
                unqUserPks = list(set(qtyObjsUsersPks+invObjsUsersPks))
                print unqUserPks
                for i in unqUserPks:
                    userObj = User.objects.get(pk=int(i))
                    flName = userObj.first_name + '_' + str(userObj.pk)
                    toRet[flName] = {'deliveredCount':0,'ongoingCount':0,'cod':0,'card':0,'total':0}

                    userqtyObjs = qtyObjs.filter(orderBy=userObj)
                    userinvObjs = invObjs.filter(orderBy=userObj)

                    deliveredCount = userqtyObjs.filter(status='delivered').count() + userinvObjs.filter(status='Completed').count()
                    ongoingCount = userqtyObjs.filter(status='outForDelivery').count() + userinvObjs.filter(status='outForDelivery').count()

                    toRet[flName]['deliveredCount'] = deliveredCount
                    toRet[flName]['ongoingCount'] = ongoingCount

                    qtyMCodAmount = userqtyObjs.filter(modeOfPayment='cash').aggregate(tot=Sum('paidAmount'))['tot']
                    qtyMCodAmount = qtyMCodAmount if qtyMCodAmount else 0
                    qtyMCardAmount = userqtyObjs.filter(modeOfPayment='card').aggregate(tot=Sum('paidAmount'))['tot']
                    qtyMCardAmount = qtyMCardAmount if qtyMCardAmount else 0

                    invMCodAmount = userinvObjs.filter(modeOfPayment='cash').aggregate(tot=Sum('amountRecieved'))['tot']
                    invMCodAmount = invMCodAmount if invMCodAmount else 0
                    invMCardAmount = userinvObjs.filter(modeOfPayment='card').aggregate(tot=Sum('amountRecieved'))['tot']
                    invMCardAmount = invMCardAmount if invMCardAmount else 0

                    cod = qtyMCodAmount + invMCodAmount
                    card = qtyMCardAmount + invMCardAmount
                    total = cod + card
                    if total == 0 and ongoingCount == 0:
                        del toRet[flName]
                    else:
                        toRet[flName]['cod'] = cod
                        toRet[flName]['card'] = card
                        toRet[flName]['total'] = total
                        toRet[flName]['name'] = flName
                        deliveryDownloadData.append([flName,deliveredCount,ongoingCount,cod,card,total])
                if 'download' in request.GET:
                    return ExcelResponse(deliveryDownloadData, 'deliveryReport_'+str(dt),'Delivery Details')

        return Response(toRet,status = status.HTTP_200_OK)


class SendShippingErrorStatus(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.IsAuthenticated ,)
    def post(self , request , format = None):
        username = request.data['username']
        firstName = request.data['firstName']
        lastName = request.data['lastName']
        address = request.data['address']
        billingAddress = request.data['billingAddress']
        email = request.data['email']
        phone = request.data['phone']
        ctx = {
            'username':username,
            'firstName':firstName,
            'lastName':lastName,
            'address':address,
            'email':email,
            'phone':phone,
            'billingAddress':billingAddress
        }
        email_body = get_template('app.ecommerce.sendShippingErrorStatus.html').render(ctx)
        email_subject = "Error in calculating shipping charges"
        email_to = []
        for i in globalSettings.G_ADMIN:
            email_to.append(str(i))
        email_cc = []
        email_bcc = []
        send_email(email_body,email_to,email_subject,email_cc,email_bcc,'html')
        return Response({"errorStatusSent":True}, status = status.HTTP_200_OK)
